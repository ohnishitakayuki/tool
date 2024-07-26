import os
import sys
import numpy as np
import pandas as pd
from pathlib import Path
from datetime import datetime

import openpyxl

sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
from getdf.convert_lreg import ConvertLreg
from getdf.pos_correct import PosCorrect


class PosQcLreg:
    v_str = {0: 'nocorr', 1: 'rot', 2: '1st'}

    # Field設定
    field_num = 5
    field_size = 2.64

    # Excel関係
    p_excel = Path(os.path.dirname(__file__) + '/../../excel_template/posQC/Local Lreg STR CH240nm format.xlsx')
    excel_start_row = 5
    excel_start_column = 2

    def __init__(self, p1, p2):
        # 補正番号をリスト化
        p1 = Path(p1)
        p2 = Path(p2)

        self.sw = [0, 1, 2]
        df1 = self._convert_df(p1)
        df2 = self._convert_df(p2)

        # csv化のためのコピー
        self.df1_org = df1.copy()
        self.df2_org = df2.copy()

        # 測定時間。ファイル取得時間から受け取る
        meas_start_1st = datetime.fromtimestamp(p1.stat().st_mtime)
        meas_start_2nd = datetime.fromtimestamp(p2.stat().st_mtime)
        v_meas = {'meas_start_1st': meas_start_1st, 'meas_start_2nd': meas_start_2nd,}

        # # 補正
        v_tmp = self._calc_rep(df1, df2)
        v_coef1 = self._calc_coef(df1, 'data1')
        v_coef2 = self._calc_coef(df2, 'data2')

        # # 辞書結合
        v = v_meas | v_tmp | v_coef1 | v_coef2

        for k in v:
            setattr(self, k, v[k])

    def _convert_df(self, p):
        # strかpathlibかわからんのでとりあえず変換。
        p = Path(p)
        if not(p.exists()):
            raise FileNotFoundError('ファイルがありません。')
        if p.stat().st_size == 0:
            raise FileNotFoundError('ファイルが空です。')

        # 変換
        c = ConvertLreg(p)
        df = c.df()

        return df

    def _arrange_data(self, df):
        f_num = self.field_num
        f_size = self.field_size
        df['fieldX'] = df.apply(lambda x: int((x['Field']-1) / f_num) * f_size - f_size * 2, axis=1) * 1000
        df['fieldY'] = df.apply(lambda x: int((x['Field']-1) % f_num) * f_size - f_size * 2, axis=1) * 1000
        df['X'] = df['X'] - df['fieldX']
        df['Y'] = df['Y'] - df['fieldY']
        return df

    def _arrange_sl(self, df1, df2):
        df2 = df2.rename(columns={'X': 'X_2nd', 'Y': 'Y_2nd'})
        df = pd.merge(df1, df2[['Site', 'Field', 'X_2nd', 'Y_2nd']], how='inner', on=['Site', 'Field'])
        df['X'] = (df['X'] + df['X_2nd']) / 2
        df['Y'] = (df['Y'] + df['Y_2nd']) / 2
        df = df.drop(['X_2nd', 'Y_2nd'], axis=1)
        df['num'] = range(1, len(df.index) + 1)
        f_num = self.field_num
        f_size = self.field_size
        df['fieldX'] = df.apply(lambda x: int((x['num']-1) / f_num) * f_size - f_size * 2, axis=1) * 1000
        df['fieldY'] = df.apply(lambda x: int((x['num']-1) % f_num) * f_size - f_size * 2, axis=1) * 1000
        df['X'] = df['X'] - df['fieldX']
        df['Y'] = df['Y'] - df['fieldY']
        df = df.rename(columns={'X':'X_sl', 'Y':'Y_sl'})
        df = df.drop('num', axis=1)
        return df

    def _calc_rep(self, df1_org, df2_org):
        # 深いコピーにして元df変更防止
        df1 = df1_org.copy()
        df2 = df2_org.copy()
        p1 = PosCorrect(df1)
        p2 = PosCorrect(df2)
        v = {}

        # 3s計算
        for sw in self.sw:
            df1_tmp = p1.correct_map(sw=sw)
            df2_tmp = p2.correct_map(sw=sw)
            df1_tmp = df1_tmp.rename(columns={'X': 'X_1st', 'Y': 'Y_1st'})
            df2_tmp = df2_tmp.rename(columns={'X': 'X_2nd', 'Y': 'Y_2nd'})
            df2_tmp = df2_tmp[['designX', 'designY', 'X_2nd', 'Y_2nd']]
            df = pd.merge(df1_tmp, df2_tmp, how='inner', on=['designX', 'designY'])
            df['diffX'] = df['X_1st'] - df['X_2nd']
            df['diffY'] = df['Y_1st'] - df['Y_2nd']
            x_3s = df['diffX'].std() * 3 / np.sqrt(2)
            y_3s = df['diffY'].std() * 3 / np.sqrt(2)
            v_tmp = {f'x_3s_{self.v_str[sw]}': x_3s, f'y_3s_{self.v_str[sw]}': y_3s}
            v.update(v_tmp)

        return v

    def _calc_coef(self, df, name):
        a0 = df['X'].mean()
        b0 = df['Y'].mean()
        a1, _ = np.polyfit(df['designX'], df['X'], 1)
        a2, _ = np.polyfit(df['designY'], df['X'], 1)
        b1, _ = np.polyfit(df['designX'], df['Y'], 1)
        b2, _ = np.polyfit(df['designY'], df['Y'], 1)
        rot = (a2 - b1) / 2
        v = {f'a0_{name}': a0, f'b0_{name}': b0,
             f'a1_{name}': a1, f'b1_{name}': b1,
             f'a2_{name}': a2, f'b2_{name}': b2,
             f'rot_{name}': rot}
        return v

# Excel処理部
    def to_excel(self, p_save):
        wb = openpyxl.load_workbook(self.p_excel)
        ws = wb['Data']

        # データフレーム処理部
        df1 = self.df1_org[['X', 'Y']]
        df2 = self.df2_org[['X', 'Y']]
        self._write_excel(df1, ws)
        self._write_excel(df2, ws, offset_columns=2)

        # 測定時間処理部
        ws['h5'] = self.meas_start_1st
        ws['h6'] = self.meas_start_2nd

        # 保存
        wb.save(p_save)

    # データフレームをExcelに書き込むときに使用
    def _write_excel(self, df, ws, offset_columns=0):
        for j in range(len(df)):
            for i in range(len(df.columns)):
                ws.cell(row=self.excel_start_row+j, column=self.excel_start_column+offset_columns+i,
                        value=df.iloc[j, i])


class PosQcLreg120nm(PosQcLreg):
    p_excel = Path(os.path.dirname(__file__) + '/../../excel_template/posQC/Local Lreg STR CH120nm format.xlsx')

    def __init__(self, p1, p2):
        super().__init__(p1, p2)


if __name__ == '__main__':
    p1 = Path('../../test/test_data/data_pos_qc_lreg/QC_Local_CH240_V03-000.010.lms.LocalRegistration.M.0.8.8._._.lreg')
    p2 = Path('../../test/test_data/data_pos_qc_lreg/QC_Local_CH240_V03-000.011.lms.LocalRegistration.M.0.8.8._._.lreg')
    c = PosQcLreg(p1, p2)
    c.to_excel('../../../result/Local Lreg STR result.xlsx')
    print(vars(c))
    # c.to_excel(p_save)

