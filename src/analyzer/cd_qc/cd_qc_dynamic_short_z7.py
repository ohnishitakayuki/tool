import os
import sys
import numpy as np
import pandas as pd
from pathlib import Path

import openpyxl

from pprint import pprint

sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
from getdf.convert_at import ConvertAt
from getdf.convert_emu import ConvertEmu


class CdQcDynamicShortZ7:
    # Excel関係
    p_excel = Path(os.path.dirname(__file__) + '/../../excel_template/cdQC/Dynamic Short Z7 format.xlsx')
    excel_start_row = 3
    excel_start_column = 2

    def __init__(self, p, plate_type=''):
        df = self._convert_df(p)

        # 測定時間
        meas_start = df['meas_date'].min()
        meas_end = df['meas_date'].max()
        meas_time = meas_end - meas_start
        v_meas = {'meas_time': meas_time, 'meas_start': meas_start, 'meas_end': meas_end}

        # 値
        v_value = self._calc_rep_dynamic_short(df)

        # plate type
        self.plate_type = plate_type
        v_plate = {'plate_type': self.plate_type}

        v = v_meas | v_value
        for k in v:
            setattr(self, k, v[k])

        # columnとrowをオブジェクト変数化
        self.column_num = df['die_x'][0]
        self.row_num = df['die_y'][0]

        # Excel用の変換
        self.df_raw = df

    def _convert_df(self, p):
        # strかpathlibかわからんのでとりあえず変換。
        p = Path(p)
        if not p:
            raise FileNotFoundError('ファイルがありません。')
        if p.stat().st_size == 0:
            raise FileNotFoundError('ファイルが空です。')

        # 変換
        # Result~はEMU
        if 'Result' in p.name:
            c = ConvertEmu()
        else:
            c = ConvertAt()
        df = df = c.df(p)

        if df.empty:
            raise FileNotFoundError('ファイルが空です。')

        return df

    def _calc_rep_dynamic_short(self, df_org):
        df = df_org.copy()
        v = {}

        # df作成
        repeat_time = int(len(df) / 6)
        list_category = ['iso_space', 'iso_space', 'ls', 'ls', 'iso_line', 'iso_line'] * repeat_time
        df['category'] = list_category
        list_design_size = [110] * 6 + [200] * 6 + [400] * 6 + [750] * 6
        repeat_time_size = int(len(df) / len(list_design_size))
        list_design_size = list_design_size * repeat_time_size
        df['design_size'] = list_design_size
        repeat_time_chip = 3
        list_chip = [1] * int(len(df)/3) + [2] * int(len(df)/3) + [3] * int(len(df)/3)
        df['chip'] = list_chip

        df1 = df[['meas_date', 'design_pos_x', 'design_pos_y', 'rotation', 'category', 'cd1', 'design_size', 'chip']]
        df1 = df1.rename(columns={'label1': 'label', 'cd1': 'cd'})
        df1['category'] = df1['category'].str.replace('ls', 'ls_space')
        df2 = df[['meas_date', 'design_pos_x', 'design_pos_y', 'rotation', 'category', 'cd2', 'design_size', 'chip']]
        df2 = df2.rename(columns={'label2': 'label', 'cd2': 'cd'})
        df2 = df2.dropna()
        df2['category'] = df2['category'] + '_line'
        df_all = pd.concat([df1, df2]).reset_index().drop('index', axis=1)
        df_all['design_size'] = df_all['design_size'].astype(int)

        c_rotations = df_all['rotation'].unique()
        c_categories = df_all['category'].unique()
        c_design_sizes = df_all['design_size'].unique()
        c_chips = df_all['chip'].unique()

        j = 0
        a_all = 0
        df_rep = pd.DataFrame()
        list_slope = []

        for c_rotation in c_rotations:
            for c_category in c_categories:
                for c_design_size in c_design_sizes:
                    for c_chip in c_chips:
                        df_tmp = df_all[(df_all['rotation'] == c_rotation) & (df_all['category'] == c_category) &
                                 (df_all['design_size'] == c_design_size) & (df_all['chip'] == c_chip)]
                        df_tmp = df_tmp.sort_values('meas_date').reset_index()
                        # 平均除去
                        df_tmp['cd_remove_ave'] = df_tmp['cd'] - df_tmp['cd'].mean()

                        # slope除去
                        x = df_tmp.index.values
                        y = df_tmp['cd'].values
                        a, b = np.polyfit(x, y, 1)
                        list_slope.append(abs(a))
                        df_tmp['cd_remove_1st'] = df_tmp['cd'] - a * df_tmp.index
                        df_tmp['cd_remove_1st'] = df_tmp['cd_remove_1st'] - df_tmp['cd_remove_1st'].mean()
                        df_rep = pd.concat([df_rep, df_tmp])

        df_rep = df_rep.drop(['meas_date', 'index'], axis=1)
        df_cd = df_rep.groupby(['design_pos_x', 'design_pos_y', 'rotation', 'category', 'design_size', 'chip']).mean().reset_index()
        df_3s = df_rep.groupby(['design_pos_x', 'design_pos_y', 'rotation', 'category', 'design_size', 'chip']).std().reset_index()

        df_3s['cd_remove_ave'] = df_3s['cd_remove_ave'] * 3
        df_3s['cd_remove_1st'] = df_3s['cd_remove_1st'] * 3
        df_3s = df_3s.rename(columns={'cd_remove_ave': 'rep_3s_nocorr', 'cd_remove_1st': 'rep_3s_1st'})
        df_3s = df_3s.drop('cd', axis=1)
        df_ave = pd.merge(df_cd, df_3s, how='inner', on=['design_pos_x', 'design_pos_y', 'rotation', 'category', 'design_size', 'chip'])
        self.df_ave = df_ave

        rep_3s_raw = df_ave['rep_3s_nocorr'].mean()
        rep_3s_1st = df_ave['rep_3s_1st'].mean()
        slope_ave = np.mean(list_slope)

        # 全体
        v['rep_3s_raw'] = rep_3s_raw
        v['rep_3s_1st'] = rep_3s_1st
        v['slope_average'] = slope_ave

        # チップ
        for c_chip in c_chips:
            df_tmp = df_ave[df_ave['chip']==c_chip]
            rep_nocorr = df_tmp['rep_3s_nocorr'].mean()
            rep_1st = df_tmp['rep_3s_1st'].mean()
            v[f'rep_3s_raw_chip{c_chip}'] = rep_nocorr
            v[f'rep_3s_1st_chip{c_chip}'] = rep_1st

        # サイズ
        for c_design_size in c_design_sizes:
            df_tmp = df_ave[df_ave['design_size']==c_design_size]
            rep_nocorr = df_tmp['rep_3s_nocorr'].mean()
            rep_1st = df_tmp['rep_3s_1st'].mean()
            v[f'rep_3s_raw_{c_design_size}nm'] = rep_nocorr
            v[f'rep_3s_1st_{c_design_size}nm'] = rep_1st

        # カテゴリー
        for c_category in c_categories:
            df_tmp = df_ave[df_ave['category'] == c_category]
            rep_nocorr = df_tmp['rep_3s_nocorr'].mean()
            rep_1st = df_tmp['rep_3s_1st'].mean()
            v[f'rep_3s_raw_{c_category}'] = rep_nocorr
            v[f'rep_3s_1st_{c_category}'] = rep_1st

        # XY
        for c_rotation in c_rotations:
            df_tmp = df_ave[df_ave['rotation'] == c_rotation]
            rep_nocorr = df_tmp['rep_3s_nocorr'].mean()
            rep_1st = df_tmp['rep_3s_1st'].mean()
            if c_rotation == 0:
                s_rotation = 'x'
            elif c_rotation == 90:
                s_rotation = 'y'
            v[f'rep_3s_raw_{s_rotation}'] = rep_nocorr
            v[f'rep_3s_1st_{s_rotation}'] = rep_1st

        return v


# Excel処理部
    def to_excel(self, p_save):
        wb = openpyxl.load_workbook(self.p_excel)
        ws = wb['Data']

        # データフレーム処理部
        df = self.df_raw[['cd1', 'cd2', ]]

        self._write_excel(df, ws)

        # 測定時間処理部
        ws['i3'] = self.meas_start
        ws['i4'] = self.meas_start
        ws['i5'] = self.meas_end

        # 測定チップ入力
        ws['f3'] = self.column_num
        ws['f4'] = self.row_num

        # plate名
        ws['v7'] = self.plate_type

        # 保存
        wb.save(p_save)
    #
    # データフレームをExcelに書き込むときに使用
    def _write_excel(self, df, ws, offset_columns=0):
        for j in range(len(df)):
            for i in range(len(df.columns)):
                ws.cell(row=self.excel_start_row+j, column=self.excel_start_column+offset_columns+i,
                        value=df.iloc[j, i])

if __name__ == '__main__':
    p = Path('../../../data/DAT-Dynamic_Short_V01_6_6/ResultMain.CSV')
    c = CdQcDynamicShortZ7(p, plate_type='No.1')
    # p_excel = Path('../../excel_template/Global STR format.xlsx')
    p_save = Path('../../../result/dynamic_short_z7.xlsx')
    pprint(vars(c))
    # c.to_excel(p_save)
