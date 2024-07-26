import os
import sys
import numpy as np
import pandas as pd
from pathlib import Path

import openpyxl

sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
from getdf.convert_lms import ConvertLms
from getdf.convert_prx import ConvertPrx
from getdf.pos_correct import PosCorrect


class PosQcLpos:
    v_str = {0: 'nocorr', 1: 'rot', 2: '1st'}

    # Field設定
    field_num = 5
    field_size = 2.64

    # Excel関係
    p_excel = Path(os.path.dirname(__file__) + '/../../excel_template/posQC/Local STR format.xlsx')
    excel_start_row = 5
    excel_start_column = 2

    def __init__(self, p1, p2, p1_sl1, p1_sl2, p2_sl1, p2_sl2):
        # 補正番号をリスト化
        self.sw = [0, 1, 2]
        df1 = self._convert_df(p1)
        df2 = self._convert_df(p2)
        df1_sl1 = self._convert_df(p1_sl1)
        df1_sl2 = self._convert_df(p1_sl2)
        df2_sl1 = self._convert_df(p2_sl1)
        df2_sl2 = self._convert_df(p2_sl2)

        # csv化のためのコピー
        self.df1_org = df1.copy()
        self.df2_org = df2.copy()
        self.df1_sl1_org = df1_sl1.copy()
        self.df1_sl2_org = df1_sl2.copy()
        self.df2_sl1_org = df2_sl1.copy()
        self.df2_sl2_org = df2_sl2.copy()

        # 測定時間
        meas_start_1st = df1['meas_date'].min()
        meas_end_1st = df1['meas_date'].max()
        meas_start_2nd = df2['meas_date'].min()
        meas_end_2nd = df2['meas_date'].max()
        meas_time_1st = meas_end_1st - meas_start_1st
        meas_time_2nd = meas_end_2nd - meas_start_2nd
        v_meas = {'meas_time_1st': meas_time_1st, 'meas_start_1st': meas_start_1st, 'meas_end_1st': meas_end_1st,
                  'meas_time_2nd': meas_time_2nd, 'meas_start_2nd': meas_start_2nd, 'meas_end_2nd': meas_end_2nd}

        # Field座標などを追加
        df1 = self._arrange_data(df1)
        df2 = self._arrange_data(df2)

        # SLは別関数。2つ送って平均化。
        df1_sl = self._arrange_sl(df1_sl1, df1_sl2)
        df2_sl = self._arrange_sl(df2_sl1, df2_sl2)

        # SL補正
        df1 = pd.merge(df1, df1_sl[['X_sl', 'Y_sl', 'fieldX', 'fieldY']], how='left', on=['fieldX', 'fieldY'])
        df2 = pd.merge(df2, df2_sl[['X_sl', 'Y_sl', 'fieldX', 'fieldY']], how='left', on=['fieldX', 'fieldY'])

        df1['X'] = df1['X'] - df1['X_sl']
        df1['Y'] = df1['Y'] - df1['Y_sl']
        df2['X'] = df2['X'] - df2['X_sl']
        df2['Y'] = df2['Y'] - df2['Y_sl']
        df1['designX'] = df1['designX'] + df1['fieldX']
        df1['designY'] = df1['designY'] + df1['fieldY']
        df2['designX'] = df2['designX'] + df2['fieldX']
        df2['designY'] = df2['designY'] + df2['fieldY']

        # 補正
        v_tmp = self._calc_rep(df1, df2)
        v_coef1 = self._calc_coef(df1, 'data1')
        v_coef2 = self._calc_coef(df2, 'data2')

        # 辞書結合
        v = v_meas | v_tmp | v_coef1 | v_coef2

        for k in v:
            setattr(self, k, v[k])

    def _convert_df(self, p):
        # strかpathlibかわからんのでとりあえず変換。
        p = Path(p)
        if not(p.exists()):
            raise FileNotFoundError('ファイルがありません。')
        if p.stat().st_size == 0:
            raise FileNotFoundError('ファイルが空です。')

        # 変換。拡張子でファイルタイプ確認。
        if p.suffix == '.lms':
            c = ConvertLms(p)
        elif p.suffix == '.prx':
            c = ConvertPrx(p)
        else:
            raise FileNotFoundError('ファイルが対応している拡張子ではありません。')

        df = c.df(calc='lpos')

        return df

    def _arrange_data(self, df):
        f_num = self.field_num
        f_size = self.field_size
        df['fieldX'] = df.apply(lambda x: int((x['Field']-1) / f_num) * f_size - f_size * 2, axis=1) * 1000
        df['fieldY'] = df.apply(lambda x: int((x['Field']-1) % f_num) * f_size - f_size * 2, axis=1) * 1000
        df['X'] = df['X'] - df['fieldX']
        df['Y'] = df['Y'] - df['fieldY']
        return df

    def _arrange_sl(self, df1, df2):
        df2 = df2.rename(columns={'X': 'X_2nd', 'Y': 'Y_2nd'})
        df = pd.merge(df1, df2[['Site', 'Field', 'X_2nd', 'Y_2nd']], how='inner', on=['Site', 'Field'])
        df['X'] = (df['X'] + df['X_2nd']) / 2
        df['Y'] = (df['Y'] + df['Y_2nd']) / 2
        df = df.drop(['X_2nd', 'Y_2nd'], axis=1)
        df['num'] = range(1, len(df.index) + 1)
        f_num = self.field_num
        f_size = self.field_size
        df['fieldX'] = df.apply(lambda x: int((x['num']-1) / f_num) * f_size - f_size * 2, axis=1) * 1000
        df['fieldY'] = df.apply(lambda x: int((x['num']-1) % f_num) * f_size - f_size * 2, axis=1) * 1000
        df['X'] = df['X'] - df['fieldX']
        df['Y'] = df['Y'] - df['fieldY']
        df = df.rename(columns={'X':'X_sl', 'Y':'Y_sl'})
        df = df.drop('num', axis=1)
        return df

    def _calc_rep(self, df1_org, df2_org):
        # 深いコピーにして元df変更防止
        df1 = df1_org.copy()
        df2 = df2_org.copy()
        p1 = PosCorrect(df1)
        p2 = PosCorrect(df2)
        v = {}

        # 3s計算
        for sw in self.sw:
            df1_tmp = p1.correct_map(sw=sw)
            df2_tmp = p2.correct_map(sw=sw)
            df1_tmp = df1_tmp.rename(columns={'X': 'X_1st', 'Y': 'Y_1st'})
            df2_tmp = df2_tmp.rename(columns={'X': 'X_2nd', 'Y': 'Y_2nd'})
            df2_tmp = df2_tmp[['designX', 'designY', 'X_2nd', 'Y_2nd']]
            df = pd.merge(df1_tmp, df2_tmp, how='inner', on=['designX', 'designY'])
            df['diffX'] = df['X_1st'] - df['X_2nd']
            df['diffY'] = df['Y_1st'] - df['Y_2nd']
            x_3s = df['diffX'].std() * 3 / np.sqrt(2)
            y_3s = df['diffY'].std() * 3 / np.sqrt(2)
            v_tmp = {f'x_3s_{self.v_str[sw]}': x_3s, f'y_3s_{self.v_str[sw]}': y_3s}
            v.update(v_tmp)

        return v

    def _calc_coef(self, df, name):
        a0 = df['X'].mean()
        b0 = df['Y'].mean()
        a1, _ = np.polyfit(df['designX'], df['X'], 1)
        a2, _ = np.polyfit(df['designY'], df['X'], 1)
        b1, _ = np.polyfit(df['designX'], df['Y'], 1)
        b2, _ = np.polyfit(df['designY'], df['Y'], 1)
        rot = (a2 - b1) / 2
        v = {f'a0_{name}': a0, f'b0_{name}': b0,
             f'a1_{name}': a1, f'b1_{name}': b1,
             f'a2_{name}': a2, f'b2_{name}': b2,
             f'rot_{name}': rot}
        return v

# Excel処理部
    def to_excel(self, p_save):
        wb = openpyxl.load_workbook(self.p_excel)
        ws = wb['Data']

        # データフレーム処理部
        df1 = self.df1_org[['X', 'Y']]
        df2 = self.df2_org[['X', 'Y']]
        df1_sl1 = self.df1_sl1_org[['X', 'Y']]
        df1_sl2 = self.df1_sl2_org[['X', 'Y']]
        df2_sl1 = self.df2_sl1_org[['X', 'Y']]
        df2_sl2 = self.df2_sl2_org[['X', 'Y']]
        self._write_excel(df1, ws)
        self._write_excel(df2, ws, offset_columns=2)
        self._write_excel(df1_sl1, ws, offset_columns=4)
        self._write_excel(df1_sl2, ws, offset_columns=6)
        self._write_excel(df2_sl1, ws, offset_columns=8)
        self._write_excel(df2_sl2, ws, offset_columns=10)

        # 測定時間処理部
        ws['p5'] = self.meas_start_1st
        ws['p6'] = self.meas_start_2nd
        ws['q5'] = self.meas_end_1st
        ws['q6'] = self.meas_end_2nd

        # 保存
        wb.save(p_save)

    # データフレームをExcelに書き込むときに使用
    def _write_excel(self, df, ws, offset_columns=0):
        for j in range(len(df)):
            for i in range(len(df.columns)):
                ws.cell(row=self.excel_start_row+j, column=self.excel_start_column+offset_columns+i,
                        value=df.iloc[j, i])


if __name__ == '__main__':
    p_lms1 = Path('../../test/test_data/data_pos_qc_lpos/QC_Local_V03-000.000.lms')
    p_lms2 = Path('../../test/test_data/data_pos_qc_lpos/QC_Local_V03-000.010.lms')
    p_sl0 = Path('../../test/test_data/data_pos_qc_lpos/QC_Local_SL_V03-000.000.lms')
    p_sl1 = Path('../../test/test_data/data_pos_qc_lpos/QC_Local_SL_V03-000.010.lms')
    p_sl2 = Path('../../test/test_data/data_pos_qc_lpos/QC_Local_SL_V03-000.020.lms')
    c = PosQcLpos(p_lms1, p_lms2, p_sl0, p_sl1, p_sl1, p_sl2)
    p_save = Path('../../../result/Local STR result.xlsx')
    # c.to_excel(p_save)
    print(vars(c))
