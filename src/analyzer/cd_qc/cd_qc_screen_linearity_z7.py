import os
import sys
import numpy as np
import pandas as pd
from pathlib import Path

import openpyxl

from pprint import pprint

sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
from getdf.convert_at import ConvertAt
from getdf.convert_emu import ConvertEmu


class CdQcScreenLinearityZ7:
    # Excel関係
    p_excel = Path(os.path.dirname(__file__) + '/../../excel_template/cdQC/Screen Linearity z7 format.xlsx')
    excel_start_row = 3
    excel_start_column = 2

    def __init__(self, p, plate_type=''):
        df = self._convert_df(p)

        # 測定時間
        meas_start = df['meas_date'].min()
        meas_end = df['meas_date'].max()
        meas_time = meas_end - meas_start
        v_meas = {'meas_time': meas_time, 'meas_start': meas_start, 'meas_end': meas_end}

        # 値
        v_value = self._calc_screen_linearity(df)

        # plate type
        self.plate_type = plate_type
        v_plate = {'plate_type': self.plate_type}

        # # 結合してオブジェクト変数化
        v = v_meas | v_value | v_plate

        for k in v:
            setattr(self, k, v[k])

        # columnとrowをオブジェクト変数化
        self.column_num = df['die_x'][0]
        self.row_num = df['die_y'][0]

        # Excel用の変換
        self.df = df

    def _convert_df(self, p):
        # strかpathlibかわからんのでとりあえず変換。
        p = Path(p)
        if not p:
            raise FileNotFoundError('ファイルがありません。')
        if p.stat().st_size == 0:
            raise FileNotFoundError('ファイルが空です。')

        # 変換
        # Result~はEMU
        if 'Result' in p.name:
            c = ConvertEmu()
        else:
            c = ConvertAt()
        df = df = c.df(p)

        if df.empty:
            raise FileNotFoundError('ファイルが空です。')

        return df

    def _calc_screen_linearity(self, df_org):
        # 深いコピーにして元df変更防止
        df = df_org.copy()
        v = {}
        repeat_time = int(len(df)/4)
        list_design_size = [220, 400, 600, 800] * repeat_time
        df['design_size'] = list_design_size

        # cd値を一つの列にする。
        df2 = pd.DataFrame()
        for i in range(1, 6):
            df_tmp = df[['design_pos_x', 'design_pos_y', 'design_size', 'rotation', f'cd{i}']].copy()
            df_tmp['roi_num'] = i
            df_tmp = df_tmp.rename(columns={f'cd{i}': 'cd'})
            df2 = pd.concat([df2, df_tmp])

        # 平均値計算
        df_ave = df2.groupby(['design_size', 'roi_num']).mean().reset_index()
        df_ave['cd_diff'] = df_ave['cd'] - df_ave['design_size']
        c_design_sizes = df2['design_size'].unique()
        df_ave2 = pd.DataFrame()
        for c_design_size in c_design_sizes:
            df_tmp = df_ave[df_ave['design_size']==c_design_size].copy()
            df_tmp['cd_diff_mean'] = df_tmp['cd_diff'].mean()
            df_ave2 = pd.concat([df_ave2, df_tmp])
        df_ave2['cd_diff_remove_mean'] = df_ave2['cd_diff'] - df_ave2['cd_diff_mean']
        df_ave3 = df_ave2.groupby('design_size').agg({'cd_diff_remove_mean': ['max', 'min']}).reset_index()
        df_ave3.columns = ['_'.join(col).strip() for col in df_ave3.columns.values]
        df_ave3['cd_range'] = df_ave3['cd_diff_remove_mean_max'] - df_ave3['cd_diff_remove_mean_min']

        self.df_ave2 = df_ave2
        self.df_ave3 = df_ave3

        # Screen Linearity値を算出して辞書化
        screen_linearity = df_ave3['cd_range'].mean()

        v_value = {'screen_linearity': screen_linearity}

        v = v_value
        return v

    def _df_columns_cd_add(self, df, add_name):
        new_columns = []
        for column in df.columns:
            if 'cd' in column or 'AFQV' in column or 'PMQV' in column:
                new_columns.append(column + '_' + add_name)
            else:
                new_columns.append(column)
        df.columns = new_columns
        return df

    def _get_value(self, df, type, name):
        # type別に代入値を変える
        if type == 'space':
            cd_1st = 'cd1_1st'
            cd_2nd = 'cd1_2nd'
        elif type == 'pitch':
            cd_1st = 'cd3_1st'
            cd_2nd = 'cd3_2nd'

        # diffを追加してオブジェクト変数にしとく。execを使うが仕方ない。
        df['cd_diff'] = df[cd_1st] - df[cd_2nd]

        exec(f"self.df_{name}_{type} = df")

        # 再現性
        rep = df['cd_diff'].std() * 3 / np.sqrt(2)

        # slope
        data_half = int(len(df) / 2)
        slope_end = 50
        df_short1 = df[:slope_end]
        df_short2 = df[data_half: data_half+slope_end]
        slope_1st, _ = np.polyfit(df_short1.index, df_short1['cd_diff'], 1)
        slope_2nd, _ = np.polyfit(df_short1.index, df_short2['cd_diff'], 1)
        slope_1st = slope_1st * 50
        slope_2nd = slope_2nd * 50

        # diffのmean 2nd - 1stなのでマイナス1をかける
        diff_mean = df['cd_diff'].mean() * -1

        # cd値解析
        cd_mean_1st = df[cd_1st].mean()
        cd_mean_2nd = df[cd_2nd].mean()
        cd_3s_1st = df[cd_1st].std() * 3
        cd_3s_2nd = df[cd_2nd].std() * 3
        cd_max_1st = df[cd_1st].max()
        cd_max_2nd = df[cd_2nd].max()
        cd_min_1st = df[cd_1st].min()
        cd_min_2nd = df[cd_2nd].min()
        cd_range_1st = cd_max_1st - cd_min_1st
        cd_range_2nd = cd_max_2nd - cd_min_2nd


        v = {f'rep_3s_{name}_{type}': rep,
             f'slope_1st_line_{name}_{type}': slope_1st, f'slope_2nd_line_{name}_{type}': slope_2nd,
             f'diff_mean_{name}_{type}': diff_mean,
             f'cd_mean_1st_{name}_{type}': cd_mean_1st, f'cd_mean_2nd_{name}_{type}': cd_mean_2nd,
             f'cd_3s_1st_{name}_{type}': cd_3s_1st, f'cd_3s_2nd_{name}_{type}': cd_3s_2nd,
             f'cd_max_1st_{name}_{type}': cd_max_1st, f'cd_max_2nd_{name}_{type}': cd_max_2nd,
             f'cd_min_1st_{name}_{type}': cd_min_1st, f'cd_min_2nd_{name}_{type}': cd_min_2nd,
             f'cd_range_1st_{name}_{type}': cd_range_1st, f'cd_range_2nd_{name}_{type}': cd_range_2nd,
             }

        return v

# Excel処理部
    def to_excel(self, p_save):
        wb = openpyxl.load_workbook(self.p_excel)
        ws = wb['Data']

        # データフレーム処理部
        df = self.df[['cd1', 'cd2', 'cd3', 'cd4', 'cd5',]]

        self._write_excel(df, ws)

        # 測定時間処理部
        ws['l3'] = self.meas_start
        ws['l4'] = self.meas_start
        ws['l5'] = self.meas_end

        # 測定チップ入力
        ws['i3'] = self.column_num
        ws['i4'] = self.row_num

        # plate名
        ws['i8'] = self.plate_type

        # 保存
        wb.save(p_save)
    #
    # データフレームをExcelに書き込むときに使用
    def _write_excel(self, df, ws, offset_columns=0):
        for j in range(len(df)):
            for i in range(len(df.columns)):
                ws.cell(row=self.excel_start_row+j, column=self.excel_start_column+offset_columns+i,
                        value=df.iloc[j, i])

if __name__ == '__main__':
    p = Path('../../../data/DAT-Screen_Linearity_V01/ResultMain.CSV')
    c = CdQcScreenLinearityZ7(p, plate_type='No.1')
    # p_excel = Path('../../excel_template/Global STR format.xlsx')
    p_save = Path('../../../result/screen linearity.xlsx')
    pprint(vars(c))
    c.to_excel(p_save)
