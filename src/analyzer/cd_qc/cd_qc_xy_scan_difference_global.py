import os
import sys
import numpy as np
import pandas as pd
from pathlib import Path

import openpyxl

from pprint import pprint

sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
from getdf.convert_at import ConvertAt


class CdQcXyScanDifferenceGlobal:
    # Excel関係
    p_excel = Path(os.path.dirname(__file__) + '/../../excel_template/cdQC/XY Scan difference global format.xlsx')
    excel_start_row = 4
    excel_start_column = 2

    def __init__(self, p_0deg, p_270deg, plate_type=''):
        df_0deg = self._convert_df(p_0deg)
        df_270deg = self._convert_df(p_270deg)

        # 測定時間
        meas_start_0deg = df_0deg['meas_date'].min()
        meas_end_0deg = df_0deg['meas_date'].max()
        meas_start_270deg = df_270deg['meas_date'].min()
        meas_end_270deg = df_270deg['meas_date'].max()
        meas_time_0deg = meas_end_0deg - meas_start_0deg
        meas_time_270deg = meas_end_270deg - meas_start_270deg
        v_meas = {'meas_time': meas_time_0deg, 'meas_start': meas_start_0deg, 'meas_end': meas_end_0deg,
                  'meas_time_0deg': meas_time_0deg, 'meas_start_0deg': meas_start_0deg, 'meas_end_0deg': meas_end_0deg,
                  'meas_time_270deg': meas_time_270deg, 'meas_start_270deg': meas_start_270deg,
                  'meas_end_270deg': meas_end_270deg,
                  }

        # 値
        v_value = self._calc_xy_scan_diff(df_0deg, df_270deg)

        # plate type
        self.plate_type = plate_type
        v_plate = {'plate_type': self.plate_type}

        # 結合してオブジェクト変数化
        v = v_meas | v_value | v_plate
        for k in v:
            setattr(self, k, v[k])

        # Excel用の変換
        self.df_0deg = df_0deg
        self.df_270deg = df_270deg

    def _convert_df(self, p):
        # strかpathlibかわからんのでとりあえず変換。
        p = Path(p)
        if not p:
            raise FileNotFoundError('ファイルがありません。')
        if p.stat().st_size == 0:
            raise FileNotFoundError('ファイルが空です。')

        # 変換
        c = ConvertAt()
        df = df = c.df(p)

        if df.empty:
            raise FileNotFoundError('ファイルが空です。')

        return df

    def _calc_xy_scan_diff(self, df_0deg_org, df_270deg_org):
        # 深いコピーにして元df変更防止
        df_0deg = df_0deg_org.copy()
        df_270deg = df_270deg_org.copy()

        # df_270degの座標を回転
        df_270deg['design_pos_tmp'] = df_270deg['design_pos_x']
        df_270deg['design_pos_x'] = df_270deg['design_pos_y']
        df_270deg['design_pos_y'] = -df_270deg['design_pos_tmp']
        df_270deg = df_270deg.drop('design_pos_tmp', axis=1)

        # 列を絞ってからマージ
        df_0deg = df_0deg[['design_pos_x', 'design_pos_y', 'die_x', 'die_y', 'rotation', 'cd1', 'cd3']]
        df_270deg = df_270deg[['design_pos_x', 'design_pos_y', 'cd1', 'cd3']]
        df_0deg = self._df_columns_cd_add(df_0deg, '0deg')
        df_270deg = self._df_columns_cd_add(df_270deg, '270deg')
        df = pd.merge(df_0deg, df_270deg, how='inner', on=['design_pos_x', 'design_pos_y'])

        # 平均して方向合わせてから差分算出
        df_ave = df.groupby(['die_x', 'die_y', 'rotation']).mean().reset_index()
        df_ave_hor = df_ave[df_ave['rotation'] == 270].copy()
        df_ave_ver = df_ave[df_ave['rotation'] == 0].copy()

        df_ave_hor['cd1_diff'] = df_ave_hor['cd1_270deg'] - df_ave_hor['cd1_0deg']
        df_ave_hor['cd3_diff'] = df_ave_hor['cd3_270deg'] - df_ave_hor['cd3_0deg']
        df_ave_ver['cd1_diff'] = df_ave_ver['cd1_0deg'] - df_ave_ver['cd1_270deg']
        df_ave_ver['cd3_diff'] = df_ave_ver['cd3_0deg'] - df_ave_ver['cd3_270deg']

        # XY値算出
        df_ave_hor_center = df_ave_hor[(df_ave_hor['die_x'] == 6) & (df_ave_hor['die_y'] == 6)].reset_index()
        df_ave_ver_center = df_ave_ver[(df_ave_ver['die_x'] == 6) & (df_ave_ver['die_y'] == 6)].reset_index()
        center_hor_space = df_ave_hor_center['cd1_diff'][0]
        center_ver_space = df_ave_ver_center['cd1_diff'][0]
        global_hor_space = df_ave_hor['cd1_diff'].abs().max()
        global_ver_space = df_ave_ver['cd1_diff'].abs().max()
        center_hor_pitch = df_ave_hor_center['cd3_diff'][0]
        center_ver_pitch = df_ave_ver_center['cd3_diff'][0]
        global_hor_pitch = df_ave_hor['cd3_diff'].abs().max()
        global_ver_pitch = df_ave_ver['cd3_diff'].abs().max()

        v_value = {'center_hor_space': center_hor_space, 'center_ver_space': center_ver_space,
                'global_hor_space': global_hor_space, 'global_ver_space': global_ver_space,
                'center_hor_pitch': center_hor_pitch, 'center_ver_pitch': center_ver_pitch,
                'global_hor_pitch': global_hor_pitch, 'global_ver_pitch': global_ver_pitch,
                }

        # mean値算出
        cd_mean_hor_0deg_space = df_ave_hor['cd1_0deg'].mean()
        cd_mean_hor_270deg_space = df_ave_hor['cd1_270deg'].mean()
        cd_mean_ver_0deg_space = df_ave_ver['cd1_0deg'].mean()
        cd_mean_ver_270deg_space = df_ave_ver['cd1_270deg'].mean()
        cd_mean_hor_0deg_pitch = df_ave_hor['cd3_0deg'].mean()
        cd_mean_hor_270deg_pitch = df_ave_hor['cd3_270deg'].mean()
        cd_mean_ver_0deg_pitch = df_ave_ver['cd3_0deg'].mean()
        cd_mean_ver_270deg_pitch = df_ave_ver['cd3_270deg'].mean()

        v_mean = {'cd_mean_hor_0deg_space': cd_mean_hor_0deg_space,
                  'cd_mean_hor_270deg_space': cd_mean_hor_270deg_space,
                  'cd_mean_ver_0deg_space': cd_mean_ver_0deg_space,
                  'cd_mean_ver_270deg_space': cd_mean_ver_270deg_space,
                  'cd_mean_hor_0deg_pitch': cd_mean_hor_0deg_pitch,
                  'cd_mean_hor_270deg_pitch': cd_mean_hor_270deg_pitch,
                  'cd_mean_ver_0deg_pitch': cd_mean_ver_0deg_pitch,
                  'cd_mean_ver_270deg_pitch': cd_mean_ver_270deg_pitch,
                  }

        # xy difference算出
        df_xy_hor = df_ave_hor[['die_x', 'die_y', 'cd1_0deg', 'cd3_0deg', 'cd1_270deg', 'cd3_270deg']].copy()
        df_xy_ver = df_ave_ver[['die_x', 'die_y', 'cd1_0deg', 'cd3_0deg', 'cd1_270deg', 'cd3_270deg']].copy()
        df_xy_hor = df_xy_hor.rename(columns={'cd1_0deg': 'cd1_0deg_hor', 'cd3_0deg': 'cd3_0deg_hor',
                                              'cd1_270deg': 'cd1_270deg_hor', 'cd3_270deg': 'cd3_270deg_hor', })
        df_xy_ver = df_xy_ver.rename(columns={'cd1_0deg': 'cd1_0deg_ver', 'cd3_0deg': 'cd3_0deg_ver',
                                              'cd1_270deg': 'cd1_270deg_ver', 'cd3_270deg': 'cd3_270deg_ver', })
        df_xy = pd.merge(df_xy_hor, df_xy_ver, how='inner', on=['die_x', 'die_y'])
        df_xy['cd1_xy_0deg'] = df_xy['cd1_0deg_ver'] - df_xy['cd1_0deg_hor']
        df_xy['cd1_xy_270deg'] = df_xy['cd1_270deg_ver'] - df_xy['cd1_270deg_hor']
        df_xy['cd3_xy_0deg'] = df_xy['cd3_0deg_ver'] - df_xy['cd3_0deg_hor']
        df_xy['cd3_xy_270deg'] = df_xy['cd3_270deg_ver'] - df_xy['cd3_270deg_hor']
        xy_diff_0deg_space = df_xy['cd1_xy_0deg'].mean()
        xy_diff_270deg_space = df_xy['cd1_xy_270deg'].mean()
        xy_diff_0deg_pitch = df_xy['cd3_xy_0deg'].mean()
        xy_diff_270deg_pitch = df_xy['cd3_xy_270deg'].mean()

        v_xy = {'xy_diff_0deg_space': xy_diff_0deg_space,
                'xy_diff_270deg_space': xy_diff_270deg_space,
                'xy_diff_0deg_pitch': xy_diff_0deg_pitch,
                'xy_diff_270deg_pitch': xy_diff_270deg_pitch,
                }

        # XY range
        range_hor_space = df_ave_hor['cd1_diff'].max() - df_ave_hor['cd1_diff'].min()
        range_ver_space = df_ave_ver['cd1_diff'].max() - df_ave_ver['cd1_diff'].min()
        range_hor_pitch = df_ave_hor['cd3_diff'].max() - df_ave_hor['cd3_diff'].min()
        range_ver_pitch  = df_ave_ver['cd3_diff'].max() - df_ave_ver['cd3_diff'].min()
        v_xy_range = {'range_hor_space': range_hor_space, 'range_ver_space': range_ver_space,
                      'range_hor_pitch': range_hor_pitch, 'range_ver_pitch': range_ver_pitch}

        df_ave_hor = df_ave_hor.sort_values(['die_y', 'die_x'])
        for index, row in df_ave_hor.iterrows():
            die_x = int(row['die_x'])
            die_y = int(row['die_y'])
            v_xy_range[f'xy_scan_diff_hor_space_{die_x}_{die_y}'] = row['cd1_diff']
            v_xy_range[f'xy_scan_diff_hor_pitch_{die_x}_{die_y}'] = row['cd3_diff']

        df_ave_ver = df_ave_ver.sort_values(['die_y', 'die_x'])
        for index, row in df_ave_ver.iterrows():
            die_x = int(row['die_x'])
            die_y = int(row['die_y'])
            v_xy_range[f'xy_scan_diff_ver_space_{die_x}_{die_y}'] = row['cd1_diff']
            v_xy_range[f'xy_scan_diff_ver_pitch_{die_x}_{die_y}'] = row['cd3_diff']

        # 辞書結合
        v = v_value | v_xy | v_mean | v_xy | v_xy_range

        return v

    def _df_columns_cd_add(self, df, add_name):
        new_columns = []
        for column in df.columns:
            if 'cd' in column or 'AFQV' in column or 'PMQV' in column:
                new_columns.append(column + '_' + add_name)
            else:
                new_columns.append(column)
        df.columns = new_columns
        return df

    def _calc_cd(self, df_org, df_ref_org):
        # 深いコピーにして元df変更防止
        df = df_org.copy()
        df_ref = df_ref_org.copy()
        v = {}

        # 平均化処理
        df_ave = df[['die_x', 'die_y', 'rotation', 'design_size1', 'design_size2', 'cd1', 'cd2']].groupby([
            'die_x', 'die_y', 'rotation', 'design_size1']).mean().reset_index()
        df_ref_ave = df_ref[['die_x', 'die_y', 'rotation', 'design_size1', 'design_size2', 'cd1', 'cd2']].groupby(
            ['die_x', 'die_y', 'rotation', 'design_size1']).mean().reset_index()
        df_ref_ave = df_ref_ave.rename(columns={'cd1': 'cd1_ref', 'cd2': 'cd2_ref'})
        df_all = pd.merge(df_ave, df_ref_ave, how='inner', on=['die_x', 'die_y', 'rotation', 'design_size1'])
        df_all['cd_diff1'] = df_all['cd1'] - df_all['cd1_ref']
        df_all['cd_diff2'] = df_all['cd2'] - df_all['cd2_ref']

        # 行の情報を取得して辞書に追加
        for index, row in df_all.iterrows():
            p_size = int(row['design_size1'])
            if row['rotation'] == 270:
                direction = 'hor'
            elif row['rotation'] == 0:
                direction = 'ver'
            v[f'{direction}_{p_size}nm_space'] = row['cd_diff1']
            v[f'{direction}_{p_size}nm_pitch'] = row['cd_diff2']

        return v

# Excel処理部
    def to_excel(self, p_save):
        wb = openpyxl.load_workbook(self.p_excel)
        ws = wb['Data']

        # データフレーム処理部
        df_0deg = self.df_0deg[['cd1', 'cd3']]
        df_270deg = self.df_270deg[['cd1', 'cd3']]

        self._write_excel(df_0deg, ws)
        self._write_excel(df_270deg, ws, offset_columns=2)

        # 測定時間処理部
        ws['h3'] = self.meas_start_0deg
        ws['h4'] = self.meas_start_0deg
        ws['h5'] = self.meas_end_0deg
        ws['i3'] = self.meas_start_270deg
        ws['i4'] = self.meas_start_270deg
        ws['i5'] = self.meas_end_270deg

        # plate名
        ws['h7'] = self.plate_type

        # 保存
        wb.save(p_save)

    # データフレームをExcelに書き込むときに使用
    def _write_excel(self, df, ws, offset_columns=0):
        for j in range(len(df)):
            for i in range(len(df.columns)):
                ws.cell(row=self.excel_start_row+j, column=self.excel_start_column+offset_columns+i,
                        value=df.iloc[j, i])

if __name__ == '__main__':
    p_0 = Path('../../test/test_data/data_cd_qc_xy_scan_difference_global/XY scan difference global R0_20231111205226.csv')
    p_270 = Path('../../test/test_data/data_cd_qc_xy_scan_difference_global/XY scan difference global R270_20231111213606.csv')
    c = CdQcXyScanDifferenceGlobal(p_0, p_270, plate_type='No.1')
    p_save = Path('../../../result/xy_scan_difference.xlsx')
    pprint(vars(c))

    # c.to_excel(p_save)
