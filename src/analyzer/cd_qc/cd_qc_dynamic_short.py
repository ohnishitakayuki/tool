import os
import sys
import numpy as np
import pandas as pd
from pathlib import Path

import openpyxl

from pprint import pprint

sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
from getdf.convert_at import ConvertAt


class CdQcDynamicShort:
    # Excel関係
    p_excel = Path(os.path.dirname(__file__) + '/../../excel_template/cdQC/Dynamic Short format.xlsx')
    excel_start_row = 3
    excel_start_column = 2

    def __init__(self, p, plate_type=''):
        df = self._convert_df(p)

        # 測定時間
        meas_start = df['meas_date'].min()
        meas_end = df['meas_date'].max()
        meas_time = meas_end - meas_start
        v_meas = {'meas_time': meas_time, 'meas_start': meas_start, 'meas_end': meas_end}

        # 値
        v_value = self._calc_rep_dynamic_short(df)

        # plate type
        self.plate_type = plate_type
        v_plate = {'plate_type': self.plate_type}

        v = v_meas | v_value
        for k in v:
            setattr(self, k, v[k])

        # columnとrowをオブジェクト変数化
        self.column_num = df['die_x'][0]
        self.row_num = df['die_y'][0]

        # Excel用の変換
        self.df_raw = df

    def _convert_df(self, p):
        # strかpathlibかわからんのでとりあえず変換。
        p = Path(p)
        if not p:
            raise FileNotFoundError('ファイルがありません。')
        if p.stat().st_size == 0:
            raise FileNotFoundError('ファイルが空です。')

        # 変換
        c = ConvertAt()
        df = df = c.df(p)

        if df.empty:
            raise FileNotFoundError('ファイルが空です。')

        return df

    def _calc_rep_dynamic_short(self, df_org):
        df = df_org.copy()

        v = {}

        # df作成
        repeat_time = int(len(df) / 6)
        list_category = ['iso_space', 'iso_space', 'iso_line', 'iso_line', 'ls', 'ls'] * repeat_time
        df['category'] = list_category

        df1 = df[['meas_date', 'design_pos_x', 'design_pos_y', 'rotation', 'category',
                  'design_pos_x_ROI1', 'design_pos_y_ROI1', 'label1', 'cd1', 'design_size1']]
        df1 = df1.rename(columns={'design_pos_x_ROI1': 'design_pos_x_ROI', 'design_pos_y_ROI1': 'design_pos_y_ROI',
                                  'label1': 'label', 'cd1': 'cd', 'design_size1': 'design_size'})
        df1['category'] = df1['category'].str.replace('ls', 'ls_line')
        df2 = df[['meas_date', 'design_pos_x', 'design_pos_y', 'rotation', 'category',
                  'design_pos_x_ROI2', 'design_pos_y_ROI2', 'label2', 'cd2', 'design_size2']]
        df2 = df2.rename(columns={'design_pos_x_ROI2': 'design_pos_x_ROI', 'design_pos_y_ROI2': 'design_pos_y_ROI',
                                  'label2': 'label', 'cd2': 'cd', 'design_size2': 'design_size'})
        df2 = df2.dropna()
        df2['category'] = df2['category'] + '_space'
        df_all = pd.concat([df1, df2])
        df_all['design_size'] = df_all['design_size'].astype(int)

        c_rotations = df_all['rotation'].unique()
        c_categories = df_all['category'].unique()
        c_design_sizes = df_all['design_size'].unique()

        j = 0
        a_all = 0
        df_rep = pd.DataFrame()

        v_cd = {}
        for c_rotation in c_rotations:
            for c_category in c_categories:
                for c_design_size in c_design_sizes:
                    df_tmp = df_all[(df_all['rotation'] == c_rotation) & (df_all['category'] == c_category) &
                             (df_all['design_size'] == c_design_size)]
                    df_unique_pos = df_tmp[['design_pos_x_ROI', 'design_pos_y_ROI']].drop_duplicates()
                    chip_num = 0
                    for i, row in df_unique_pos.iterrows():
                        chip_num = chip_num + 1
                        pos_x = row['design_pos_x_ROI']
                        pos_y = row['design_pos_y_ROI']
                        df_tmp2 = df_tmp[(df_tmp['design_pos_x_ROI'] == pos_x) & (df_tmp['design_pos_y_ROI'] == pos_y)]
                        df_tmp2 = df_tmp2.sort_values('meas_date').reset_index()
                        df_tmp2['chip_num'] = chip_num
                        # 平均除去
                        df_tmp2['cd_remove_ave'] = df_tmp2['cd'] - df_tmp2['cd'].mean()

                        # slope除去
                        x = df_tmp2.index.values
                        y = df_tmp2['cd'].values
                        a, b = np.polyfit(x, y, 1)
                        df_tmp2['cd_remove_1st'] = df_tmp2['cd'] - a * df_tmp2.index
                        df_tmp2['cd_remove_1st'] = df_tmp2['cd_remove_1st'] - df_tmp2['cd_remove_1st'].mean()
                        df_rep = pd.concat([df_rep, df_tmp2])
                        j = j + 1
                        a_all = a_all + abs(a)
                        # cd mean取得
                        cd_mean = df_tmp2['cd'].mean()
                        if c_rotation == 0:
                            name_xy = 'x'
                        elif c_rotation == 270:
                            name_xy = 'y'
                        v_cd[f'cd_mean_{name_xy}_{c_category}_{c_design_size}_chip{chip_num}'] = cd_mean

        rep_3s_raw = df_rep['cd_remove_ave'].std() * 3
        rep_3s_1st = df_rep['cd_remove_1st'].std() * 3

        v['rep_3s_raw'] = rep_3s_raw
        v['rep_3s_1st'] = rep_3s_1st
        v['slope_average'] = a_all / j

        # chipの3sを取得
        c_chip_nums = df_rep['chip_num'].unique()
        for c_chip_num in c_chip_nums:
            df_tmp3 = df_rep[df_rep['chip_num'] == c_chip_num]
            rep_3s_raw_chip = df_tmp3['cd_remove_ave'].std() * 3
            rep_3s_1st_chip = df_tmp3['cd_remove_1st'].std() * 3
            v[f'rep_3s_raw_chip_{c_chip_num}'] = rep_3s_raw_chip
            v[f'rep_3s_1st_chip_{c_chip_num}'] = rep_3s_1st_chip

        # design sizeの3sを取得
        for c_design_size in c_design_sizes:
            df_tmp3 = df_rep[df_rep['design_size'] == c_design_size]
            rep_3s_raw_size = df_tmp3['cd_remove_ave'].std() * 3
            rep_3s_1st_size = df_tmp3['cd_remove_1st'].std() * 3
            v[f'rep_3s_raw_size_{c_design_size}'] = rep_3s_raw_size
            v[f'rep_3s_1st_size_{c_design_size}'] = rep_3s_1st_size

        # categoryの3sを取得
        for c_category in c_categories:
            df_tmp3 = df_rep[df_rep['category'] == c_category]
            rep_3s_raw_category = df_tmp3['cd_remove_ave'].std() * 3
            rep_3s_1st_category = df_tmp3['cd_remove_1st'].std() * 3
            v[f'rep_3s_raw_category_{c_category}'] = rep_3s_raw_category
            v[f'rep_3s_1st_category_{c_category}'] = rep_3s_1st_category

        # xyの3sを取得
        for c_rotation in c_rotations:
            df_tmp3 = df_rep[df_rep['rotation'] == c_rotation]
            rep_3s_raw_xy = df_tmp3['cd_remove_ave'].std() * 3
            rep_3s_1st_xy = df_tmp3['cd_remove_1st'].std() * 3
            if c_rotation == 0:
                name_xy = 'x'
            elif c_rotation == 270:
                name_xy = 'y'
            v[f'rep_3s_raw_rotation_{name_xy}'] = rep_3s_raw_xy
            v[f'rep_3s_1st_rotation_{name_xy}'] = rep_3s_1st_xy

        v = v | v_cd

        return v


# Excel処理部
    def to_excel(self, p_save):
        wb = openpyxl.load_workbook(self.p_excel)
        ws = wb['Data']

        # データフレーム処理部
        df = self.df_raw[['cd1', 'cd2', ]]

        self._write_excel(df, ws)

        # 測定時間処理部
        ws['i3'] = self.meas_start
        ws['i4'] = self.meas_start
        ws['i5'] = self.meas_end

        # 測定チップ入力
        ws['f3'] = self.column_num
        ws['f4'] = self.row_num

        # plate名
        ws['v7'] = self.plate_type

        # 保存
        wb.save(p_save)
    #
    # データフレームをExcelに書き込むときに使用
    def _write_excel(self, df, ws, offset_columns=0):
        for j in range(len(df)):
            for i in range(len(df.columns)):
                ws.cell(row=self.excel_start_row+j, column=self.excel_start_column+offset_columns+i,
                        value=df.iloc[j, i])

if __name__ == '__main__':
    p = Path('../../test/test_data/data_cd_qc_dynamic_short/Dynamic Short AT_20221118133357.csv')
    c = CdQcDynamicShort(p, plate_type='No.1')
    # p_excel = Path('../../excel_template/Global STR format.xlsx')
    p_save = Path('../../../result/dynamic_short.xlsx')
    # pprint(vars(c))
    list_key = []
    for k in vars(c):
        list_key.append(k)
    pprint(list_key)
    # c.to_excel(p_save)
