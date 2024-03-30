import os
import sys
import numpy as np
import pandas as pd
from pathlib import Path

import openpyxl

from pprint import pprint

sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
from getdf.convert_at import ConvertAt


class CdQcMagnificationError:
    # 初期設定
    mag_ref = 50000

    # Excel関係
    p_excel = Path(os.path.dirname(__file__) + '/../../excel_template/cdQC/Magnification Error format.xlsx')
    excel_start_row = 2
    excel_start_column = 2

    def __init__(self, p, plate_type=''):
        df = self._convert_df(p)

        # 測定時間
        meas_start = df['meas_date'].min()
        meas_end = df['meas_date'].max()
        meas_time = meas_end - meas_start
        v_meas = {'meas_time': meas_time, 'meas_start': meas_start, 'meas_end': meas_end}

        # 値
        v_value = self._calc_mag(df)

        # plate type
        self.plate_type = plate_type
        v_plate = {'plate_type': self.plate_type}

        # # 結合してオブジェクト変数化
        v = v_meas | v_value | v_plate

        for k in v:
            setattr(self, k, v[k])

        # columnとrowをオブジェクト変数化
        self.column_num = df['die_x'][0]
        self.row_num = df['die_y'][0]

        # Excel用の変換
        self.df = df

    def _convert_df(self, p):
        # strかpathlibかわからんのでとりあえず変換。
        p = Path(p)
        if not p:
            raise FileNotFoundError('ファイルがありません。')
        if p.stat().st_size == 0:
            raise FileNotFoundError('ファイルが空です。')

        # 変換
        c = ConvertAt()
        df = df = c.df(p)

        if df.empty:
            raise FileNotFoundError('ファイルが空です。')

        return df

    def _calc_mag(self, df_org):
        # 深いコピーにして元df変更防止
        df = df_org.copy()
        v = {}

        # 倍率の列を取り出す。
        df2 = df['scan'].str.split(':', expand=True)
        df2 = df2.rename(columns={1: 'magnification'})
        df2 = df2[['magnification']].astype(int)
        df = pd.concat([df, df2], axis=1)

        # Magnification値算出
        df_hor = df[df['rotation'] == 270]
        df_ver = df[df['rotation'] == 0]

        df_hor = df_hor[['magnification', 'cd1']]
        df_hor_ave = df_hor.groupby('magnification').mean().reset_index()
        cd_ref = df_hor_ave[df_hor_ave['magnification'] == self.mag_ref].reset_index().loc[0, 'cd1']
        df_hor_ave['cd_diff'] = df_hor_ave['cd1'] - cd_ref
        mag_error_hor = df_hor_ave['cd_diff'].max() - df_hor_ave['cd_diff'].min()

        df_ver = df_ver[['magnification', 'cd1']]
        df_ver_ave = df_ver.groupby('magnification').mean().reset_index()
        cd_ref = df_ver_ave[df_ver_ave['magnification'] == self.mag_ref].reset_index().loc[0, 'cd1']
        df_ver_ave['cd_diff'] = df_ver_ave['cd1'] - cd_ref
        mag_error_ver = df_ver_ave['cd_diff'].max() - df_ver_ave['cd_diff'].min()

        v_value = {'mag_error_hor': mag_error_hor, 'mag_error_ver': mag_error_ver}

        v = v_value
        return v

# Excel処理部
    def to_excel(self, p_save):
        wb = openpyxl.load_workbook(self.p_excel)
        ws = wb['Data']

        # データフレーム処理部
        df = self.df[['cd1',]]

        self._write_excel(df, ws)

        # 測定時間処理部
        ws['h3'] = self.meas_start
        ws['h4'] = self.meas_start
        ws['h5'] = self.meas_end

        # 測定チップ入力
        ws['e3'] = self.column_num
        ws['e4'] = self.row_num

        # plate名
        ws['e8'] = self.plate_type

        # 保存
        wb.save(p_save)
    #
    # データフレームをExcelに書き込むときに使用
    def _write_excel(self, df, ws, offset_columns=0):
        for j in range(len(df)):
            for i in range(len(df.columns)):
                ws.cell(row=self.excel_start_row+j, column=self.excel_start_column+offset_columns+i,
                        value=df.iloc[j, i])

if __name__ == '__main__':
    p = Path('../../test/test_data/data_cd_qc_magnification_error/Magnification Error_20231116001254.csv')
    c = CdQcMagnificationError(p, plate_type='No.1')
    # p_excel = Path('../../excel_template/Global STR format.xlsx')
    p_save = Path('../../../result/magnification error.xlsx')
    pprint(vars(c))
    c.to_excel(p_save)
