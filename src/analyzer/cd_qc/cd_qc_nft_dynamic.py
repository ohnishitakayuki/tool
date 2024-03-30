import os
import sys
import numpy as np
import pandas as pd
from pathlib import Path

import openpyxl

from pprint import pprint

sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
from getdf.convert_at import ConvertAt


class CdQcNftDynamic:
    # Excel関係
    p_excel = Path(os.path.dirname(__file__) + '/../../excel_template/cdQC/NFT Dynamic format.xlsx')
    excel_start_row = 3
    excel_start_column = 2

    def __init__(self, p, plate_type=''):
        df = self._convert_df(p)

        # 測定時間
        meas_start = df['meas_date'].min()
        meas_end = df['meas_date'].max()
        meas_time = meas_end - meas_start
        v_meas = {'meas_time': meas_time, 'meas_start': meas_start, 'meas_end': meas_end}

        # 値
        v_value = self._calc_rep_hv2(df)

        # plate type
        self.plate_type = plate_type
        v_plate = {'plate_type': self.plate_type}

        # 結合してオブジェクト変数化
        v = v_meas | v_value | v_plate
        for k in v:
            setattr(self, k, v[k])

        # columnとrowをオブジェクト変数化
        self.column_num = df['die_x'][0]
        self.row_num = df['die_y'][0]

        # Excel用の変換
        self.df_raw = df

    def _convert_df(self, p):
        # strかpathlibかわからんのでとりあえず変換。
        p = Path(p)
        if not p:
            raise FileNotFoundError('ファイルがありません。')
        if p.stat().st_size == 0:
            raise FileNotFoundError('ファイルが空です。')

        # 変換
        c = ConvertAt()
        df = df = c.df(p)

        if df.empty:
            raise FileNotFoundError('ファイルが空です。')

        return df

    def _calc_rep_hv2(self, df_org):
        # 深いコピーにして元df変更防止
        df = df_org.copy()
        v = {}
        meas_len = int(len(df) / 4)
        df_h1 = df.iloc[0: meas_len]
        df_h2 = df.iloc[meas_len: meas_len*2]
        df_v1 = df.iloc[meas_len * 2: meas_len * 3]
        df_v2 = df.iloc[meas_len * 3: meas_len * 4]

        # df内、cdがついているものに1st, 2ndをつける。
        df_h1 = self._df_columns_cd_add(df_h1, '1st')
        df_h2 = self._df_columns_cd_add(df_h2, '2nd')
        df_v1 = self._df_columns_cd_add(df_v1, '1st')
        df_v2 = self._df_columns_cd_add(df_v2, '2nd')

        df_h = pd.merge(df_h1, df_h2, how='inner', on=['design_pos_x', 'design_pos_y'])
        df_v = pd.merge(df_v1, df_v2, how='inner', on=['design_pos_x', 'design_pos_y'])

        v_hor_space = self._get_value(df_h, 'space', 'hor')
        v_ver_space = self._get_value(df_v, 'space', 'ver')
        v_hor_pitch = self._get_value(df_h, 'pitch', 'hor')
        v_ver_pitch = self._get_value(df_v, 'pitch', 'ver')

        # AFQV, PMQV値を取得
        afqv_1st_hor = df_h['AFQV_1st'].mean()
        afqv_2nd_hor = df_h['AFQV_2nd'].mean()
        afqv_1st_ver = df_v['AFQV_1st'].mean()
        afqv_2nd_ver = df_v['AFQV_2nd'].mean()
        pmqv_1st_hor = df_h['PMQV_1st'].mean()
        pmqv_2nd_hor = df_h['PMQV_2nd'].mean()
        pmqv_1st_ver = df_v['PMQV_1st'].mean()
        pmqv_2nd_ver = df_v['PMQV_2nd'].mean()
        white_band_1st_hor = df_h['cd11_1st'].mean() - df_h['cd1_1st'].mean()
        white_band_2nd_hor = df_h['cd11_2nd'].mean() - df_h['cd1_2nd'].mean()
        white_band_1st_ver = df_v['cd11_1st'].mean() - df_v['cd1_1st'].mean()
        white_band_2nd_ver = df_v['cd11_2nd'].mean() - df_v['cd1_2nd'].mean()

        v_value = {'afqv_1st_hor': afqv_1st_hor, 'afqv_2nd_hor': afqv_2nd_hor,
                   'afqv_1st_ver': afqv_1st_ver, 'afqv_2nd_ver': afqv_2nd_ver,
                   'pmqv_1st_hor': pmqv_1st_hor, 'pmqv_2nd_hor': pmqv_2nd_hor,
                   'pmqv_1st_ver': pmqv_1st_ver, 'pmqv_2nd_ver': pmqv_2nd_ver,
                   'white_band_1st_hor': white_band_1st_hor, 'white_band_2nd_hor': white_band_2nd_hor,
                   'white_band_1st_ver': white_band_1st_ver, 'white_band_2nd_ver': white_band_2nd_ver,
                    }

        v = v_hor_space | v_ver_space | v_hor_pitch | v_ver_pitch | v_value
        return v

    def _df_columns_cd_add(self, df, add_name):
        new_columns = []
        for column in df.columns:
            if 'cd' in column or 'AFQV' in column or 'PMQV' in column:
                new_columns.append(column + '_' + add_name)
            else:
                new_columns.append(column)
        df.columns = new_columns
        return df

    def _get_value(self, df, type, name):
        # type別に代入値を変える
        if type == 'space':
            cd_1st = 'cd1_1st'
            cd_2nd = 'cd1_2nd'
        elif type == 'pitch':
            cd_1st = 'cd3_1st'
            cd_2nd = 'cd3_2nd'

        # diffを追加してオブジェクト変数にしとく。execを使うが仕方ない。
        df['cd_diff'] = df[cd_1st] - df[cd_2nd]

        exec(f"self.df_{name}_{type} = df")

        # 再現性
        rep = df['cd_diff'].std() * 3 / np.sqrt(2)

        # slope
        data_half = int(len(df) / 2)
        slope_end = 50
        df_short1 = df[:slope_end]
        df_short2 = df[data_half: data_half+slope_end]
        slope_1st, _ = np.polyfit(df_short1.index, df_short1['cd_diff'], 1)
        slope_2nd, _ = np.polyfit(df_short1.index, df_short2['cd_diff'], 1)
        slope_1st = slope_1st * 50
        slope_2nd = slope_2nd * 50

        # diffのmean 2nd - 1stなのでマイナス1をかける
        diff_mean = df['cd_diff'].mean() * -1

        # cd値解析
        cd_mean_1st = df[cd_1st].mean()
        cd_mean_2nd = df[cd_2nd].mean()
        cd_3s_1st = df[cd_1st].std() * 3
        cd_3s_2nd = df[cd_2nd].std() * 3
        cd_max_1st = df[cd_1st].max()
        cd_max_2nd = df[cd_2nd].max()
        cd_min_1st = df[cd_1st].min()
        cd_min_2nd = df[cd_2nd].min()
        cd_range_1st = cd_max_1st - cd_min_1st
        cd_range_2nd = cd_max_2nd - cd_min_2nd


        v = {f'rep_3s_{name}_{type}': rep,
             f'slope_1st_line_{name}_{type}': slope_1st, f'slope_2nd_line_{name}_{type}': slope_2nd,
             f'diff_mean_{name}_{type}': diff_mean,
             f'cd_mean_1st_{name}_{type}': cd_mean_1st, f'cd_mean_2nd_{name}_{type}': cd_mean_2nd,
             f'cd_3s_1st_{name}_{type}': cd_3s_1st, f'cd_3s_2nd_{name}_{type}': cd_3s_2nd,
             f'cd_max_1st_{name}_{type}': cd_max_1st, f'cd_max_2nd_{name}_{type}': cd_max_2nd,
             f'cd_min_1st_{name}_{type}': cd_min_1st, f'cd_min_2nd_{name}_{type}': cd_min_2nd,
             f'cd_range_1st_{name}_{type}': cd_range_1st, f'cd_range_2nd_{name}_{type}': cd_range_2nd,
             }

        return v

# Excel処理部
    def to_excel(self, p_save):
        wb = openpyxl.load_workbook(self.p_excel)
        ws = wb['Data']

        # データフレーム処理部
        df = self.df_raw[['cd1', 'cd2', 'cd3', 'cd4', 'cd5', 'cd6', 'cd7', 'cd8', 'cd9', 'cd10', 'cd11',
                        'design_pos_x', 'design_pos_y', 'real_pos_x', 'real_pos_y', 'real_pos_z', 'AFQV', 'PMQV']]

        self._write_excel(df, ws)

        # 測定時間処理部
        ws['y3'] = self.meas_start
        ws['y4'] = self.meas_start
        ws['y5'] = self.meas_end

        # 測定チップ入力
        ws['v3'] = self.column_num
        ws['v4'] = self.row_num

        # plate名
        ws['v8'] = self.plate_type

        # 保存
        wb.save(p_save)
    #
    # データフレームをExcelに書き込むときに使用
    def _write_excel(self, df, ws, offset_columns=0):
        for j in range(len(df)):
            for i in range(len(df.columns)):
                ws.cell(row=self.excel_start_row+j, column=self.excel_start_column+offset_columns+i,
                        value=df.iloc[j, i])

if __name__ == '__main__':
    p = Path('../../test/test_data/data_cd_qc_nft_dynamic_trend/NFT Dynamic_20231220102342.csv')
    c = CdQcNftDynamic(p, plate_type='No.1')
    # p_excel = Path('../../excel_template/Global STR format.xlsx')
    p_save = Path('../../../result/nft dynamic.xlsx')
    print(vars(c))
    c.to_excel(p_save)
