import os
import sys
import numpy as np
import pandas as pd
from pathlib import Path

import openpyxl

sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
from getdf.convert_lms import ConvertLms
from getdf.pos_correct import PosCorrect


class PosQcSl:
    v_str = {'nocorr': 'nocorr', 0: 'shift'}

    # Field設定
    field_num = 11
    field_size = 2.64

    # Excel関係
    p_excel = Path(os.path.dirname(__file__) + '/../../excel_template/posQC/Screen Linearity format.xlsx')
    excel_start_row = 4
    excel_start_column = 2


    def __init__(self, p):
        # 補正番号をリスト化
        self.sw = ['nocorr', 0]
        df = self._convert_df(p)
        self.df = df

        # 測定時間
        meas_start = df['meas_date'].min()
        meas_end = df['meas_date'].max()
        meas_time = meas_end - meas_start

        self.meas_start = meas_start
        self.meas_end = meas_end
        v_meas = {'meas_time': meas_time, 'meas_start': meas_start, 'meas_end': meas_end}

        # data
        v_tmp = self._calc_sl(df)

        # 辞書結合
        v = v_meas | v_tmp

        for k in v:
            setattr(self, k, v[k])

    def _convert_df(self, p):
        # strかpathlibかわからんのでとりあえず変換。
        p = Path(p)
        if not p:
            raise FileNotFoundError('ファイルがありません。')
        if p.stat().st_size == 0:
            raise FileNotFoundError('ファイルが空です。')

        # 変換
        c = ConvertLms(p)
        df = c.df(calc='sl')

        return df

    def _calc_sl(self, df_org):
        # 深いコピーにして元df変更防止
        df = df_org.copy()
        p = PosCorrect(df)
        v = {}

        for sw in self.sw:
            if sw == 'nocorr':
                df_tmp = df
            elif sw == 0:
                df_tmp = p.correct_map(sw=sw)
            x_3s = df_tmp['X'].std() * 3
            x_mean = df_tmp['X'].mean()
            x_max = df_tmp['X'].max()
            x_min = df_tmp['X'].min()
            x_range = x_max - x_min
            y_3s = df_tmp['Y'].std() * 3
            y_mean = df_tmp['Y'].mean()
            y_max = df_tmp['Y'].max()
            y_min = df_tmp['Y'].min()
            y_range = y_max - y_min
            v_tmp = {f'x_3s_{self.v_str[sw]}': x_3s, f'x_mean_{self.v_str[sw]}': x_mean,
                     f'x_max_{self.v_str[sw]}': x_max, f'x_min_{self.v_str[sw]}': x_min,
                     f'x_range_{self.v_str[sw]}': x_range,
                     f'y_3s_{self.v_str[sw]}': y_3s, f'y_mean_{self.v_str[sw]}': y_mean,
                     f'y_max_{self.v_str[sw]}': y_max, f'y_min_{self.v_str[sw]}': y_min,
                     f'y_range_{self.v_str[sw]}': y_range,
                     }
            v.update(v_tmp)

        return v


    def _calc_coef(self, df, name):
        a0 = df['X'].mean()
        b0 = df['Y'].mean()
        a1, _ = np.polyfit(df['designX'], df['X'], 1)
        a2, _ = np.polyfit(df['designY'], df['X'], 1)
        b1, _ = np.polyfit(df['designX'], df['Y'], 1)
        b2, _ = np.polyfit(df['designY'], df['Y'], 1)
        rot = (a2 - b1) / 2
        v = {f'a0_{name}': a0, f'b0_{name}': b0,
           f'a1_{name}': a1, f'b1_{name}': b1,
           f'a2_{name}': a2, f'b2_{name}': b2,
           f'rot_{name}': rot}
        return v

# Excel処理部
    def to_excel(self, p_save):
        wb = openpyxl.load_workbook(self.p_excel)
        ws = wb['Data']

        # データフレーム処理部
        dfw = self.df[['X', 'Y']]
        self._write_excel(dfw, ws)

        # 測定時間処理部
        ws['f3'] = self.meas_start
        ws['f4'] = self.meas_end

        # 保存
        wb.save(p_save)

    # データフレームをExcelに書き込むときに使用
    def _write_excel(self, df, ws, offset_columns=0):
        for j in range(len(df)):
            for i in range(len(df.columns)):
                ws.cell(row=self.excel_start_row+j, column=self.excel_start_column+offset_columns+i,
                        value=df.iloc[j, i])


if __name__ == '__main__':
    p = Path('../../test/test_data/data_pos_qc_sl/QC_Screen_Linearity_V03-000.010.lms')
    c = PosQcSl(p)
    p_save = Path('../../../result/Screen Linearity.xlsx')
    # c.to_excel(p_save)
    print(vars(c))