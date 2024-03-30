import os
import sys
import numpy as np
import pandas as pd
from pathlib import Path
from pprint import pprint

import openpyxl

sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
from getdf.convert_lms import ConvertLms
from getdf.pos_correct import PosCorrect


class PosQcSpc:
    v_str = {0: 'nocorr', 1: 'rot', 2: '1st'}


    # Excel関係
    p_excel = Path(os.path.dirname(__file__) + '/../../excel_template/posQC/SPC format.xlsx')
    excel_start_row = 5
    excel_start_column = 2


    def __init__(self, list_p_baseline, list_p):
        # ファイルはリストで受け取る。
        self.sw = [0, 1, 2]


        df_baseline = self._get_df(list_p_baseline)
        df_data = self._get_df(list_p)

        self.df_baseline = df_baseline
        self.df_data = df_data

        # 測定時間
        meas_start = df_data['meas_date'].min()
        meas_end = df_data['meas_date'].max()
        meas_time = meas_end - meas_start

        v_meas = {'meas_time': meas_time, 'meas_start': meas_start, 'meas_end': meas_end}

        # 係数取得
        v_tmp = self._calc_spc(df_baseline, df_data)


        # 辞書結合
        v = v_meas | v_tmp

        for k in v:
            setattr(self, k, v[k])

    def _get_df(self, list_p):
        # baselineをdfに変換する関数
        df = pd.DataFrame()
        for i, p in enumerate(list_p):
            df_tmp = self._convert_df(p)
            df_tmp['data_num'] = i + 1
            df = pd.concat([df, df_tmp])
        return df

    def _convert_df(self, p):
        # strかpathlibかわからんのでとりあえず変換。
        p = Path(p)
        print(p)
        if not p:
            raise FileNotFoundError('ファイルがありません。')
        if p.stat().st_size == 0:
            raise FileNotFoundError('ファイルが空です。')

        # 変換
        c = ConvertLms(p)
        df = c.df(calc='gpos')

        return df

    def _calc_spc(self, df_baseline_org, df_data_org):
        # 深いコピーにして元df変更防止
        df_baseline = df_baseline_org.copy()
        df_data = df_data_org.copy()

        # ベースライン平均マップ化
        df_baseline = df_baseline[['designX', 'designY', 'X1', 'Y1']]
        df_baseline_ave = df_baseline.groupby(['designX', 'designY']).mean().reset_index()
        df_baseline_ave = df_baseline_ave.rename(columns={'X1': 'X1_baseline', 'Y1': 'Y1_baseline'})

        # それぞれのマップに対して値を取得
        data_nums = df_data['data_num'].unique()
        v = {}
        for data_num in data_nums:
            df = df_data[df_data['data_num'] == data_num]
            df = pd.merge(df, df_baseline_ave, how='inner', on=['designX', 'designY'])
            df['X'] = df['X1'] - df['X1_baseline']
            df['Y'] = df['Y1'] - df['Y1_baseline']
            t = PosCorrect(df)
            for sw in self.sw:
                df2 = t.correct_map(sw=sw)
                x_spc = df2['X'].std() * 3
                y_spc = df2['Y'].std() * 3
                v_tmp = {f'x_spc_data{data_num}_{self.v_str[sw]}': x_spc,
                        f'y_spc_data{data_num}_{self.v_str[sw]}': y_spc}
                v.update(v_tmp)

        v_max = {}
        for k in v:
            for sw in self.sw:
                directions = ['x', 'y']
                for direction in directions:
                    if direction in k and self.v_str[sw] in k:
                        v_max.setdefault(f'{direction}_spc_{self.v_str[sw]}_max', []).append(v[k])
        for k in v_max:
            v_max[k] = max(v_max[k])

        v.update(v_max)
        return v

    def _calc_coef(self, df, name):
        a0 = df['X'].mean()
        b0 = df['Y'].mean()
        a1, _ = np.polyfit(df['designX'], df['X'], 1)
        a2, _ = np.polyfit(df['designY'], df['X'], 1)
        b1, _ = np.polyfit(df['designX'], df['Y'], 1)
        b2, _ = np.polyfit(df['designY'], df['Y'], 1)
        rot = (a2 - b1) / 2
        v = {f'a0_{name}': a0, f'b0_{name}': b0,
           f'a1_{name}': a1, f'b1_{name}': b1,
           f'a2_{name}': a2, f'b2_{name}': b2,
           f'rot_{name}': rot}
        return v

# Excel処理部
    def to_excel(self, p_save):
        wb = openpyxl.load_workbook(self.p_excel)
        ws = wb['Data']

        # データ取得部
        df_b_w = self._get_merge_df(self.df_baseline)
        df_d_w = self._get_merge_df(self.df_data)

        # 貼り付け部
        column_size = len(df_b_w.columns)
        self._write_excel(df_b_w, ws)
        self._write_excel(df_d_w, ws, offset_columns=column_size)

        # 測定時間処理部
        ws['br4'] = self.meas_start
        ws['bs4'] = self.meas_end


        # 保存
        wb.save(p_save)

    def _get_merge_df(self, df):
        df_tmp = df[['designX', 'designY', 'Row', 'Col', 'data_num']]
        df_tmp = df_tmp[df_tmp['data_num'] == 1]
        df_w = df_tmp[['designX', 'designY', 'Row', 'Col']]

        d_columns = df['data_num'].unique()
        for d_column in d_columns:
            df_tmp2 = df[df['data_num'] == d_column]
            df_tmp2 = df_tmp2[['designX', 'designY', 'X1', 'Y1']]
            df_tmp2 = df_tmp2.rename(columns={'X1': f'X_{d_column}', 'Y1': f'Y_{d_column}'})
            df_w = pd.merge(df_w, df_tmp2, how='inner', on=['designX', 'designY'])
        df_w = df_w.drop(['designX', 'designY', 'Row', 'Col'], axis=1)
        return df_w

    # データフレームをExcelに書き込むときに使用
    def _write_excel(self, df, ws, offset_columns=0):
        for j in range(len(df)):
            for i in range(len(df.columns)):
                ws.cell(row=self.excel_start_row+j, column=self.excel_start_column+offset_columns+i,
                        value=df.iloc[j, i])


if __name__ == '__main__':
    p_baseline = Path('../../test/test_data/data_pos_qc_spc/baseline')
    p = Path('../../test/test_data/data_pos_qc_spc/data')
    # p_save = Path('../../../result/SPC result.xlsx')
    list_p_baseline = list(p_baseline.glob('*.lms'))
    list_p = list(p.glob('*.lms'))
    c = PosQcSpc(list_p_baseline, list_p)
    # c.to_excel(p_save)
    pprint(vars(c))