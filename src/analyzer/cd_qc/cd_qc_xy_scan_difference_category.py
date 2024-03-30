import os
import sys
import numpy as np
import pandas as pd
from pathlib import Path

import openpyxl

from pprint import pprint

sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
from getdf.convert_at import ConvertAt


class CdQcXyScanDifferenceCategory:
    # Excel関係
    p_excel = Path(os.path.dirname(__file__) + '/../../excel_template/cdQC/XY scan difference category format.xlsx')
    excel_start_row = 4
    excel_start_column = 2

    def __init__(self, p_0deg, p_270deg, plate_type=''):
        df_0deg = self._convert_df(p_0deg)
        df_270deg = self._convert_df(p_270deg)

        # 測定時間
        meas_start_0deg = df_0deg['meas_date'].min()
        meas_end_0deg = df_0deg['meas_date'].max()
        meas_start_270deg = df_270deg['meas_date'].min()
        meas_end_270deg = df_270deg['meas_date'].max()
        meas_time_0deg = meas_end_0deg - meas_start_0deg
        meas_time_270deg = meas_end_270deg - meas_start_270deg
        meas_time = meas_time_0deg + meas_time_270deg
        v_meas = {'meas_time': meas_time, 'meas_start': meas_start_0deg, 'meas_end': meas_end_270deg,
                  'meas_time_0deg': meas_time_0deg, 'meas_start_0deg': meas_start_0deg, 'meas_end_0deg': meas_end_0deg,
                  'meas_time_270deg': meas_time_270deg, 'meas_start_270deg': meas_start_270deg,
                  'meas_end_270deg': meas_end_270deg,
                  }

        # 値
        v_value = self._calc_xy_scan_diff(df_0deg, df_270deg)

        # plate type
        self.plate_type = plate_type
        v_plate = {'plate_type': self.plate_type}

        # 結合してオブジェクト変数化
        v = v_meas | v_value | v_plate
        for k in v:
            setattr(self, k, v[k])

        # Excel用の変換
        self.df_0deg = df_0deg
        self.df_270deg = df_270deg

    def _convert_df(self, p):
        # strかpathlibかわからんのでとりあえず変換。
        p = Path(p)
        if not p:
            raise FileNotFoundError('ファイルがありません。')
        if p.stat().st_size == 0:
            raise FileNotFoundError('ファイルが空です。')

        # 変換
        c = ConvertAt()
        df = df = c.df(p)

        if df.empty:
            raise FileNotFoundError('ファイルが空です。')

        return df

    def _calc_xy_scan_diff(self, df_0deg_org, df_270deg_org):
        # 深いコピーにして元df変更防止
        df_0deg = df_0deg_org.copy()
        df_270deg = df_270deg_org.copy()

        # df_270degの座標を回転
        df_270deg['design_pos_tmp'] = df_270deg['design_pos_x']
        df_270deg['design_pos_x'] = df_270deg['design_pos_y']
        df_270deg['design_pos_y'] = -df_270deg['design_pos_tmp']
        df_270deg = df_270deg.drop('design_pos_tmp', axis=1)

        # 列を絞ってからマージ
        df_0deg = df_0deg[['design_pos_x', 'design_pos_y', 'design_size1', 'design_pos_y_ROI1',
                           'design_pos_y_ROI2', 'design_pos_y_ROI3', 'cd1', 'cd2', 'cd3']]
        df_270deg = df_270deg[['design_pos_x', 'design_pos_y', 'cd1', 'cd2', 'cd3']]
        df_0deg = self._df_columns_cd_add(df_0deg, '0deg')
        df_270deg = self._df_columns_cd_add(df_270deg, '270deg')
        df = pd.merge(df_0deg, df_270deg, how='inner', on=['design_pos_x', 'design_pos_y'])

        # チップ割り振り
        repeat_time = int(len(df) / 5)
        list_chip = ['chip1'] * repeat_time + ['chip2'] * repeat_time + ['chip3'] * repeat_time +\
                    ['chip4'] * repeat_time + ['chip5'] * repeat_time
        df['chip'] = list_chip

        # カテゴリー割り振り
        repeat_time_cate = int(len(df) / 15)
        list_category_iso_space = ['iso_space'] * 5
        list_category_iso_line = ['iso_line'] * 5
        list_category_ls = ['ls'] * 5
        list_category = list_category_iso_space + list_category_iso_line + list_category_ls
        list_category = list_category * repeat_time_cate
        df['category'] = list_category

        # cd2, cd3をcd1の列に移動
        df1 = df[['design_pos_x', 'design_pos_y', 'design_size1', 'chip', 'category', 'design_pos_y_ROI1',
                  'cd1_0deg', 'cd1_270deg']]
        df2 = df[['design_pos_x', 'design_pos_y', 'design_size1', 'chip', 'category', 'design_pos_y_ROI2',
                  'cd2_0deg', 'cd2_270deg']]
        df3 = df[['design_pos_x', 'design_pos_y', 'design_size1', 'chip', 'category', 'design_pos_y_ROI3',
                  'cd3_0deg', 'cd3_270deg']]
        df1 = df1.rename(columns={'design_pos_y_ROI1': 'design_pos_y_ROI', 'cd1_0deg': 'cd_0deg',
                                  'cd1_270deg': 'cd_270deg'})
        df2 = df2.rename(columns={'design_pos_y_ROI2': 'design_pos_y_ROI', 'cd2_0deg': 'cd_0deg',
                                  'cd2_270deg': 'cd_270deg'})
        df3 = df3.rename(columns={'design_pos_y_ROI3': 'design_pos_y_ROI', 'cd3_0deg': 'cd_0deg',
                                  'cd3_270deg': 'cd_270deg'})
        df2 = df2.dropna()
        df3 = df3.dropna()
        df1['category'] = df1['category'].str.replace('ls', 'ls_line')
        df2['category'] = df2['category'] + '_space'
        df3['category'] = df2['category'] + '_pitch'

        df_all = pd.concat([df1, df2])
        df_all = pd.concat([df_all, df3])

        # 平均して方向合わせてから差分算出
        df_ave = df_all.groupby(['design_size1', 'chip', 'category']).mean().reset_index()
        df_ave['cd_diff'] = df_ave['cd_0deg'] - df_ave['cd_270deg']
        df_ave['cd_diff_ave'] = df_ave['cd_diff'].abs()

        # dfのオブジェクト変数化
        self.df_ave = df_ave

        # chip平均
        df_ave_chip = df_all[['design_size1', 'category',
                              'cd_0deg', 'cd_270deg']].groupby(['design_size1', 'category']).mean().reset_index()
        df_ave_chip['cd_diff'] = df_ave_chip['cd_0deg'] - df_ave_chip['cd_270deg']
        df_ave_chip['cd_diff_ave'] = df_ave_chip['cd_diff'].abs()

        xy_ave = df_ave['cd_diff_ave'].mean()
        xy_bias = df_ave['cd_diff_ave'].max()
        xy_bias_chip = df_ave_chip['cd_diff_ave'].max()

        v_value = {'xy_ave': xy_ave, 'xy_bias': xy_bias, 'xy_bias_chip': xy_bias_chip}

        return v_value

    def _df_columns_cd_add(self, df, add_name):
        new_columns = []
        for column in df.columns:
            if 'cd' in column or 'AFQV' in column or 'PMQV' in column:
                new_columns.append(column + '_' + add_name)
            else:
                new_columns.append(column)
        df.columns = new_columns
        return df

# Excel処理部
    def to_excel(self, p_save):
        wb = openpyxl.load_workbook(self.p_excel)
        ws = wb['Data']

        # データフレーム処理部
        df_0deg = self.df_0deg[['cd1', 'cd2', 'cd3']]
        df_270deg = self.df_270deg[['cd1', 'cd2', 'cd3']]

        self._write_excel(df_0deg, ws)
        self._write_excel(df_270deg, ws, offset_columns=3)

        # 測定時間処理部
        ws['j3'] = self.meas_start_0deg
        ws['j4'] = self.meas_start_0deg
        ws['j5'] = self.meas_end_0deg
        ws['k3'] = self.meas_start_270deg
        ws['k4'] = self.meas_start_270deg
        ws['k5'] = self.meas_end_270deg

        # plate名
        ws['j7'] = self.plate_type

        # 保存
        wb.save(p_save)

    # データフレームをExcelに書き込むときに使用
    def _write_excel(self, df, ws, offset_columns=0):
        for j in range(len(df)):
            for i in range(len(df.columns)):
                ws.cell(row=self.excel_start_row+j, column=self.excel_start_column+offset_columns+i,
                        value=df.iloc[j, i])

if __name__ == '__main__':
    p_0 = Path('../../test/test_data/data_cd_qc_xy_scan_difference_category/XY scan difference category R0_20231202130241.csv')
    p_270 = Path('../../test/test_data/data_cd_qc_xy_scan_difference_category/XY scan difference category R270_20231202133239.csv')
    c = CdQcXyScanDifferenceCategory(p_0, p_270, plate_type='No.1')
    p_save = Path('../../../result/xy_scan_difference_category.xlsx')
    pprint(vars(c))

    # c.to_excel(p_save)
