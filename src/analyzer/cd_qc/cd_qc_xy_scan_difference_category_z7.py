import os
import sys
import numpy as np
import pandas as pd
from pathlib import Path

import openpyxl

from pprint import pprint

sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
from getdf.convert_at import ConvertAt
from getdf.convert_emu import ConvertEmu


class CdQcXyScanDifferenceCategoryZ7:
    # Excel関係
    p_excel = Path(os.path.dirname(__file__) + '/../../excel_template/cdQC/XY scan difference category z7 format.xlsx')
    excel_start_row = 4
    excel_start_column = 2

    def __init__(self, p_0deg, p_90deg, plate_type=''):
        df_0deg = self._convert_df(p_0deg)
        df_90deg = self._convert_df(p_90deg)

        # 測定時間
        meas_start_0deg = df_0deg['meas_date'].min()
        meas_end_0deg = df_0deg['meas_date'].max()
        meas_start_90deg = df_90deg['meas_date'].min()
        meas_end_90deg = df_90deg['meas_date'].max()
        meas_time_0deg = meas_end_0deg - meas_start_0deg
        meas_time_90deg = meas_end_90deg - meas_start_90deg
        meas_time = meas_time_0deg + meas_time_90deg
        v_meas = {'meas_time': meas_time, 'meas_start': meas_start_0deg, 'meas_end': meas_end_90deg,
                  'meas_time_0deg': meas_time_0deg, 'meas_start_0deg': meas_start_0deg, 'meas_end_0deg': meas_end_0deg,
                  'meas_time_90deg': meas_time_90deg, 'meas_start_90deg': meas_start_90deg,
                  'meas_end_90deg': meas_end_90deg,
                  }

        # 値
        v_value = self._calc_xy_scan_diff(df_0deg, df_90deg)

        # plate type
        self.plate_type = plate_type
        v_plate = {'plate_type': self.plate_type}

        # 結合してオブジェクト変数化
        v = v_meas | v_value | v_plate
        for k in v:
            setattr(self, k, v[k])

        # Excel用の変換
        self.df_0deg = df_0deg
        self.df_90deg = df_90deg

    def _convert_df(self, p):
        # strかpathlibかわからんのでとりあえず変換。
        p = Path(p)
        if not p:
            raise FileNotFoundError('ファイルがありません。')
        if p.stat().st_size == 0:
            raise FileNotFoundError('ファイルが空です。')

        # 変換
        # Result~はEMU
        if 'Result' in p.name:
            c = ConvertEmu()
        else:
            c = ConvertAt()
        df = df = c.df(p)

        if df.empty:
            raise FileNotFoundError('ファイルが空です。')

        return df

    def _calc_xy_scan_diff(self, df_0deg_org, df_90deg_org):
        # 深いコピーにして元df変更防止
        df_0deg = df_0deg_org.copy()
        df_90deg = df_90deg_org.copy()

        # df_90degの座標を回転
        df_90deg['design_pos_tmp'] = df_90deg['design_pos_x']
        df_90deg['design_pos_x'] = -df_90deg['design_pos_y']
        df_90deg['design_pos_y'] = df_90deg['design_pos_tmp']
        df_90deg = df_90deg.drop('design_pos_tmp', axis=1)

        # 列を絞ってからマージ
        df_0deg = df_0deg[['design_pos_x', 'design_pos_y', 'cd1', 'cd2', 'cd3']]
        df_90deg = df_90deg[['design_pos_x', 'design_pos_y', 'cd1', 'cd2', 'cd3']]
        df_0deg = self._df_columns_cd_add(df_0deg, '0deg')
        df_90deg = self._df_columns_cd_add(df_90deg, '90deg')
        df = pd.merge(df_0deg, df_90deg, how='inner', on=['design_pos_x', 'design_pos_y'])

        # 寸法割り振り
        repeat_time_cate = int(len(df) / 6)
        list_size_tmp = [110, 200, 300, 400, 500, 750]
        list_size = list_size_tmp * repeat_time_cate
        df['size'] = list_size

        # 寸法で平均して方向合わせてから差分算出
        df_ave = df.groupby(['size']).mean().reset_index()
        df_ave['cd1_diff'] = df_ave['cd1_0deg'] - df_ave['cd1_90deg']
        df_ave['cd2_diff'] = df_ave['cd2_0deg'] - df_ave['cd2_90deg']
        df_ave['cd3_diff'] = df_ave['cd3_0deg'] - df_ave['cd3_90deg']
        df_ave['cd1_diff_ave'] = df_ave['cd1_diff'].abs()
        df_ave['cd2_diff_ave'] = df_ave['cd2_diff'].abs()
        df_ave['cd3_diff_ave'] = df_ave['cd3_diff'].abs()

        xy_space = df_ave['cd1_diff_ave'].mean()
        xy_line = df_ave['cd2_diff_ave'].mean()
        xy_pitch = df_ave['cd3_diff_ave'].mean()

        v_value = {'xy_space': xy_space, 'xy_line': xy_line, 'xy_pitch': xy_pitch}

        return v_value

    def _df_columns_cd_add(self, df, add_name):
        new_columns = []
        for column in df.columns:
            if 'cd' in column or 'AFQV' in column or 'PMQV' in column:
                new_columns.append(column + '_' + add_name)
            else:
                new_columns.append(column)
        df.columns = new_columns
        return df

# Excel処理部
    def to_excel(self, p_save):
        wb = openpyxl.load_workbook(self.p_excel)
        ws = wb['Data']

        # データフレーム処理部
        df_0deg = self.df_0deg[['cd1', 'cd2', 'cd3']]
        df_90deg = self.df_90deg[['cd1', 'cd2', 'cd3']]

        self._write_excel(df_0deg, ws)
        self._write_excel(df_90deg, ws, offset_columns=3)

        # 測定時間処理部
        ws['j3'] = self.meas_start_0deg
        ws['j4'] = self.meas_start_0deg
        ws['j5'] = self.meas_end_0deg
        ws['k3'] = self.meas_start_90deg
        ws['k4'] = self.meas_start_90deg
        ws['k5'] = self.meas_end_90deg

        # plate名
        ws['j7'] = self.plate_type

        # 保存
        wb.save(p_save)

    # データフレームをExcelに書き込むときに使用
    def _write_excel(self, df, ws, offset_columns=0):
        for j in range(len(df)):
            for i in range(len(df.columns)):
                ws.cell(row=self.excel_start_row+j, column=self.excel_start_column+offset_columns+i,
                        value=df.iloc[j, i])

if __name__ == '__main__':
    p_0 = Path('../../../data/DAT-XY_scan_difference_category_R0_V01/ResultMain.CSV')
    p_90 = Path('../../../data/DAT-XY_scan_difference_category_R90_V01/ResultMain.CSV')
    c = CdQcXyScanDifferenceCategoryZ7(p_0, p_90, plate_type='No.1')
    p_save = Path('../../../result/xy_scan_difference_category_z7.xlsx')
    pprint(vars(c))

    c.to_excel(p_save)
