import os
import sys
import numpy as np
import pandas as pd
from pathlib import Path

import openpyxl

from pprint import pprint

sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
from getdf.convert_at import ConvertAt


class CdQcAngle:
    # Excel関係
    p_excel = Path(os.path.dirname(__file__) + '/../../excel_template/cdQC/Angle format.xlsx')
    excel_start_row = 3
    excel_start_column = 2

    # Angle計算用
    angle_type = [-90, -75, -60, -45, -30, -15, 0, 15, 30, 45, 60, 75]
    angle_scan = [270, 270, 45, 45, 45, 0, 0, 0, 315, 315, 315, 270]
    # 角度とスキャンタイプを分けている。補正で使用。
    angle_scan_type = {'p_angle': angle_type, 'scan_angle': angle_scan}
    ave_num = 25

    def __init__(self, p_0deg, p_270deg, plate_type=''):
        df_0deg = self._convert_df(p_0deg)
        df_270deg = self._convert_df(p_270deg)

        # 測定時間
        meas_start_0deg = df_0deg['meas_date'].min()
        meas_end_0deg = df_0deg['meas_date'].max()
        meas_start_270deg = df_270deg['meas_date'].min()
        meas_end_270deg = df_270deg['meas_date'].max()
        meas_time_0deg = meas_end_0deg - meas_start_0deg
        meas_time_270deg = meas_end_270deg - meas_start_270deg
        v_meas = {'meas_time': meas_time_0deg, 'meas_start': meas_start_0deg, 'meas_end': meas_end_0deg,
                  'meas_time_0deg': meas_time_0deg, 'meas_start_0deg': meas_start_0deg, 'meas_end_0deg': meas_end_0deg,
                  'meas_time_270deg': meas_time_270deg, 'meas_start_270deg': meas_start_270deg,
                  'meas_end_270deg': meas_end_270deg,
                  }

        # 値
        v_value = self._calc_angle(df_0deg, df_270deg)

        # plate type
        self.plate_type = plate_type
        v_plate = {'plate_type': self.plate_type}

        # 結合してオブジェクト変数化
        v = v_meas | v_value | v_plate
        for k in v:
            setattr(self, k, v[k])

        # Excel用の変換
        self.df_0deg = df_0deg
        self.df_270deg = df_270deg

    def _convert_df(self, p):
        # strかpathlibかわからんのでとりあえず変換。
        p = Path(p)
        if not p:
            raise FileNotFoundError('ファイルがありません。')
        if p.stat().st_size == 0:
            raise FileNotFoundError('ファイルが空です。')

        # 変換
        c = ConvertAt()
        df = df = c.df(p)

        if df.empty:
            raise FileNotFoundError('ファイルが空です。')

        return df

    def _calc_angle(self, df_0deg_org, df_270deg_org):
        # 深いコピーにして元df変更防止
        df_0deg = df_0deg_org.copy()
        df_270deg = df_270deg_org.copy()
        df_0deg = self._get_angle(df_0deg)
        df_270deg = self._get_angle(df_270deg)

        # df_270の座標を回転
        df_270deg['design_pos_x_tmp'] = df_270deg['design_pos_x']
        df_270deg['design_pos_x'] = df_270deg['design_pos_y']
        df_270deg['design_pos_y'] = -df_270deg['design_pos_x_tmp']
        df_270deg = df_270deg.drop('design_pos_x_tmp', axis=1)

        # cd名リネームしてからマージ
        df_0deg = df_0deg.rename(columns={'cd1': 'cd_0'})
        df_270deg = df_270deg.rename(columns={'cd1': 'cd_270'})
        df_270deg = df_270deg[['design_pos_x', 'design_pos_y', 'cd_270']]

        df = pd.merge(df_0deg, df_270deg, how='inner', on=['design_pos_x', 'design_pos_y'])

        # 平均化
        df_ave = df.groupby('p_angle').agg({'cd_0': 'mean', 'cd_270': 'mean'}).reset_index()
        df_ave['cd_diff'] = df_ave['cd_270'] - df_ave['cd_0']

        # スキャンごとの平均を求めて、差分を求める。
        df_ave2 = df_ave.copy()
        df_scan = pd.DataFrame(self.angle_scan_type)
        df_ave2 = pd.merge(df_ave2, df_scan, how='left', on='p_angle')
        df_scan_ave = df_ave2.groupby('scan_angle').agg({'cd_diff': 'mean'}).reset_index()
        df_scan_ave = df_scan_ave.rename(columns={'cd_diff': 'cd_diff_scan'})
        df_ave_corr = pd.merge(df_ave2, df_scan_ave, how='left', on='scan_angle')
        df_ave_corr['cd_diff_corr'] = df_ave_corr['cd_diff'] - df_ave_corr['cd_diff_scan']
        df_ave_corr = df_ave_corr[
            ['p_angle', 'scan_angle', 'cd_0', 'cd_270', 'cd_diff', 'cd_diff_scan', 'cd_diff_corr']]

        # Angle値取得
        angle_diff = df_ave_corr['cd_diff'].max() - df_ave_corr['cd_diff'].min()
        v = {'angle_diff': angle_diff }

        # Angle diffを入れる
        for angle in self.angle_type:
            angle_diff = df_ave_corr[df_ave_corr['p_angle'] == angle]['cd_diff'].iloc[0]
            angle_name = str(angle).replace('-', 'm')
            v[f'cd_diff_{angle_name}deg'] = angle_diff

        return v

    def _df_columns_cd_add(self, df, add_name):
        new_columns = []
        for column in df.columns:
            if 'cd' in column or 'AFQV' in column or 'PMQV' in column:
                new_columns.append(column + '_' + add_name)
            else:
                new_columns.append(column)
        df.columns = new_columns
        return df

    def _get_angle(self, df):
        df['p_angle'] = 0
        for i, a in enumerate(self.angle_type):
            df.loc[i * self.ave_num: (i+1) * self.ave_num, ['p_angle']] = a
        return df

# Excel処理部
    def to_excel(self, p_save):
        wb = openpyxl.load_workbook(self.p_excel)
        ws = wb['Data']

        # データフレーム処理部
        df_0deg = self.df_0deg[['cd1']]
        df_270deg = self.df_270deg[['cd1']]

        self._write_excel(df_0deg, ws)
        self._write_excel(df_270deg, ws, offset_columns=1)

        # 測定時間処理部
        ws['f3'] = self.meas_start_0deg
        ws['f4'] = self.meas_start_0deg
        ws['f5'] = self.meas_end_0deg
        ws['g3'] = self.meas_start_270deg
        ws['g4'] = self.meas_start_270deg
        ws['g5'] = self.meas_end_270deg

        # plate名
        ws['f7'] = self.plate_type

        # 保存
        wb.save(p_save)

    # データフレームをExcelに書き込むときに使用
    def _write_excel(self, df, ws, offset_columns=0):
        for j in range(len(df)):
            for i in range(len(df.columns)):
                ws.cell(row=self.excel_start_row+j, column=self.excel_start_column+offset_columns+i,
                        value=df.iloc[j, i])


# 6ROI版。E3650で使用。
class CdQcAngle6Roi(CdQcAngle):
    # Excel関係
    p_excel = Path(os.path.dirname(__file__) + '/../../excel_template/cdQC/Angle 6ROI format.xlsx')
    excel_start_row = 4
    excel_start_column = 2

    def __init__(self, p_0deg, p_270deg, plate_type=''):
        super().__init__(p_0deg, p_270deg, plate_type)

    # Excel処理部。6ROIなのでその分増やす。
    def to_excel(self, p_save):
        wb = openpyxl.load_workbook(self.p_excel)
        ws = wb['Data']

        # データフレーム処理部
        df_0deg = self.df_0deg[['cd1', 'cd2', 'cd3', 'cd4', 'cd5', 'cd6']]
        df_270deg = self.df_270deg[['cd1', 'cd2', 'cd3', 'cd4', 'cd5', 'cd6']]

        self._write_excel(df_0deg, ws)
        self._write_excel(df_270deg, ws, offset_columns=6)

        # 測定時間処理部
        ws['p3'] = self.meas_start_0deg
        ws['p4'] = self.meas_start_0deg
        ws['p5'] = self.meas_end_0deg
        ws['q3'] = self.meas_start_270deg
        ws['q4'] = self.meas_start_270deg
        ws['q5'] = self.meas_end_270deg

        # plate名
        ws['p7'] = self.plate_type

        # 保存
        wb.save(p_save)



if __name__ == '__main__':
    # p_0 = Path('../../test/test_data/data_cd_qc_angle/Angle R0_20231229073100.csv')
    # p_270 = Path('../../test/test_data/data_cd_qc_angle/Angle R270_20231229075635.csv')
    # c = CdQcAngle(p_0, p_270, plate_type='No.1')
    # p_save = Path('../../../result/angle.xlsx')
    # print(vars(c))

    # c.to_excel(p_save)
    p_0 = Path('../../test/test_data/data_cd_qc_angle_6roi/Angle_R0_V01_1-8.000H.csv')
    p_270 = Path('../../test/test_data/data_cd_qc_angle_6roi/Angle_R270_V01_1-8.000H.csv')
    c = CdQcAngle6Roi(p_0, p_270, plate_type='No.1')
    p_save = Path('../../../result/angle 6roi.xlsx')
    print(vars(c))

    c.to_excel(p_save)

