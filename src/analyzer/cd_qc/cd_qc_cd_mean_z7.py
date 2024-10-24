import os
import sys
import numpy as np
import pandas as pd
from pathlib import Path

import openpyxl

from pprint import pprint

sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
from getdf.convert_at import ConvertAt
from getdf.convert_emu import ConvertEmu


class CdQcCdMeanZ7:
    # Excel関係
    p_excel = Path(os.path.dirname(__file__) + '/../../excel_template/cdQC/CD mean Z7 format.xlsx')
    excel_start_row = 3
    excel_start_column = 2

    def __init__(self, p, p_ref, plate_type=''):
        df = self._convert_df(p)
        df_ref = self._convert_df(p_ref)

        # 測定時間
        meas_start = df['meas_date'].min()
        meas_end = df['meas_date'].max()
        meas_time = meas_end - meas_start
        v_meas = {'meas_time': meas_time, 'meas_start': meas_start, 'meas_end': meas_end}

        # 値
        v_value = self._calc_cd(df, df_ref)

        # plate type
        self.plate_type = plate_type
        v_plate = {'plate_type': self.plate_type}

        # 結合してオブジェクト変数化
        v = v_meas | v_value | v_plate
        for k in v:
            setattr(self, k, v[k])

        # columnとrowをオブジェクト変数化
        self.column_num = df['die_x'][0]
        self.row_num = df['die_y'][0]

        # Excel用の変換
        self.df = df
        self.df_ref = df_ref

    def _convert_df(self, p):
        # strかpathlibかわからんのでとりあえず変換。
        p = Path(p)
        if not p:
            raise FileNotFoundError('ファイルがありません。')
        if p.stat().st_size == 0:
            raise FileNotFoundError('ファイルが空です。')

        # 変換
        # Result~はEMU
        if 'Result' in p.name:
            c = ConvertEmu()
        else:
            c = ConvertAt()
        df = df = c.df(p)

        if df.empty:
            raise FileNotFoundError('ファイルが空です。')

        return df

    def _calc_cd(self, df_org, df_ref_org):
        # 深いコピーにして元df変更防止
        df = df_org.copy()
        df_ref = df_ref_org.copy()
        v = {}
        l_design1 = ([100] * 16 + [200] * 16 + [500] * 16) * 2
        l_design2 = ([200] * 16 + [400] * 16 + [1000] * 16) * 2
        l_ref_design1 = l_design1 * int((len(df_ref) / len(df)))
        l_ref_design2 = l_design2 * int((len(df_ref) / len(df)))
        df['design_size1'] = l_design1
        df['design_size2'] = l_design2
        df_ref['design_size1'] = l_ref_design1
        df_ref['design_size2'] = l_ref_design2
        # cd3→cd2へリネーム
        df = df.drop('cd2', axis=1)
        df = df.rename(columns={'cd3': 'cd2'})
        df_ref = df_ref.drop('cd2', axis=1)
        df_ref = df_ref.rename(columns={'cd3': 'cd2'})
        # 平均化処理
        df_ave = df[['die_x', 'die_y', 'rotation', 'design_size1', 'design_size2', 'cd1', 'cd2']].groupby([
            'die_x', 'die_y', 'rotation', 'design_size1']).mean().reset_index()
        df_ref_ave = df_ref[['die_x', 'die_y', 'rotation', 'design_size1', 'design_size2', 'cd1', 'cd2']].groupby(
            ['die_x', 'die_y', 'rotation', 'design_size1']).mean().reset_index()
        df_ref_ave = df_ref_ave.rename(columns={'cd1': 'cd1_ref', 'cd2': 'cd2_ref'})
        df_all = pd.merge(df_ave, df_ref_ave, how='inner', on=['die_x', 'die_y', 'rotation', 'design_size1'])
        df_all['cd_diff1'] = df_all['cd1'] - df_all['cd1_ref']
        df_all['cd_diff2'] = df_all['cd2'] - df_all['cd2_ref']

        # 行の情報を取得して辞書に追加
        df_all.to_csv('../../../t3.csv')
        for index, row in df_all.iterrows():
            p_size = int(row['design_size1'])
            if row['rotation'] == 90:
                direction = 'hor'
            elif row['rotation'] == 0:
                direction = 'ver'
            v[f'{direction}_{p_size}nm_space'] = row['cd_diff1']
            v[f'{direction}_{p_size}nm_pitch'] = row['cd_diff2']

        return v

# Excel処理部
    def to_excel(self, p_save):
        wb = openpyxl.load_workbook(self.p_excel)
        ws = wb['Data']

        # データフレーム処理部
        df = self.df[['cd1', 'cd3']]
        df_ref = self.df_ref[['cd1', 'cd3']]

        self._write_excel(df, ws)
        self._write_excel(df_ref, ws, offset_columns=4)

        # 測定時間処理部
        ws['m3'] = self.meas_start
        ws['m4'] = self.meas_start
        ws['m5'] = self.meas_end

        # 測定チップ入力
        ws['j3'] = self.column_num
        ws['j4'] = self.row_num

        # plate名
        ws['j8'] = self.plate_type

        # 保存
        wb.save(p_save)
    #
    # データフレームをExcelに書き込むときに使用
    def _write_excel(self, df, ws, offset_columns=0):
        for j in range(len(df)):
            for i in range(len(df.columns)):
                ws.cell(row=self.excel_start_row+j, column=self.excel_start_column+offset_columns+i,
                        value=df.iloc[j, i])

if __name__ == '__main__':
    p = Path('../../../data/CD mean/ResultMain.CSV')
    p_ref = Path('../../../data/CD mean all/ResultMain.CSV')
    c = CdQcCdMeanZ7(p, p_ref, plate_type='No.1')
    p_save = Path('../../../result/cd_mean.xlsx')
    print(vars(c))
    c.to_excel(p_save)
