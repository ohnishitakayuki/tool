import os
import sys
import numpy as np
import pandas as pd
from pathlib import Path

import openpyxl

sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
from getdf.convert_lms import ConvertLms
from getdf.pos_correct import PosCorrect


class PosQcGpos:
    v_str = {0: 'nocorr', 1: 'rot', 2: '1st'}

    # Excel関係
    p_excel = Path(os.path.dirname(__file__) + '/../../excel_template/posQC/Global STR format.xlsx')
    excel_start_row = 6
    excel_start_column = 2


    def __init__(self, p1, p2):
        # 補正番号をリスト化
        self.sw = [0, 1, 2]
        df1 = self._convert_df(p1)
        df2 = self._convert_df(p2)
        self.df1 = df1
        self.df2 = df2

        # 測定時間
        meas_start_1st = df1['meas_date'].min()
        meas_end_1st = df1['meas_date'].max()
        meas_start_2nd = df2['meas_date'].min()
        meas_end_2nd = df2['meas_date'].max()
        meas_time_1st = meas_end_1st - meas_start_1st
        meas_time_2nd = meas_end_2nd - meas_start_2nd
        v_meas = {'meas_time_1st': meas_time_1st, 'meas_start_1st': meas_start_1st, 'meas_end_1st': meas_end_1st,
                  'meas_time_2nd': meas_time_2nd, 'meas_start_2nd': meas_start_2nd, 'meas_end_2nd': meas_end_2nd}

        # s1
        df1_s1 = df1.rename(columns={'X1': 'X', 'Y1': 'Y'})
        df2_s1 = df2.rename(columns={'X1': 'X', 'Y1': 'Y'})
        v_tmp1 = self._calc_rep(df1_s1, df2_s1, 's1')
        v_coef1_s1 = self._calc_coef(df1_s1, 'data1_s1')
        v_coef2_s1 = self._calc_coef(df2_s1, 'data2_s1')

        # s2
        df1_s2 = df1.rename(columns={'X2': 'X', 'Y2': 'Y'})
        df2_s2 = df2.rename(columns={'X2': 'X', 'Y2': 'Y'})
        v_tmp2 = self._calc_rep(df1_s2, df2_s2, 's2')
        v_coef1_s2 = self._calc_coef(df1_s2, 'data1_s2')
        v_coef2_s2 = self._calc_coef(df2_s2, 'data2_s2')

        # 辞書結合
        v = v_meas | v_tmp1 | v_tmp2 | v_coef1_s1 | v_coef2_s1 | v_coef1_s2 | v_coef2_s2

        for k in v:
            setattr(self, k, v[k])

    def _convert_df(self, p):
        # strかpathlibかわからんのでとりあえず変換。
        p = Path(p)
        if not p:
            raise FileNotFoundError('ファイルがありません。')
        if p.stat().st_size == 0:
            raise FileNotFoundError('ファイルが空です。')

        # 変換
        c = ConvertLms(p)
        df = c.df(calc='gpos')

        return df

    def _calc_rep(self, df1_org, df2_org, name):
        # 深いコピーにして元df変更防止
        df1 = df1_org.copy()
        df2 = df2_org.copy()

        p1 = PosCorrect(df1)
        p2 = PosCorrect(df2)
        v = {}

        # 3s計算
        for sw in self.sw:
            df1_tmp = p1.correct_map(sw=sw)
            df2_tmp = p2.correct_map(sw=sw)
            df1_tmp = df1_tmp.rename(columns={'X': 'X_1st', 'Y': 'Y_1st'})
            df2_tmp = df2_tmp.rename(columns={'X': 'X_2nd', 'Y': 'Y_2nd'})
            df2_tmp = df2_tmp[['designX', 'designY', 'X_2nd', 'Y_2nd']]
            df = pd.merge(df1_tmp, df2_tmp, how='inner', on=['designX', 'designY'])

            df['diffX'] = df['X_1st'] - df['X_2nd']
            df['diffY'] = df['Y_1st'] - df['Y_2nd']
            x_3s = df['diffX'].std() * 3 / np.sqrt(2)
            y_3s = df['diffY'].std() * 3 / np.sqrt(2)
            v_tmp = {f'x_3s_{name}_{self.v_str[sw]}': x_3s, f'y_3s_{name}_{self.v_str[sw]}': y_3s}
            v.update(v_tmp)

        return v

    def _calc_coef(self, df, name):
        a0 = df['X'].mean()
        b0 = df['Y'].mean()
        a1, _ = np.polyfit(df['designX'], df['X'], 1)
        a2, _ = np.polyfit(df['designY'], df['X'], 1)
        b1, _ = np.polyfit(df['designX'], df['Y'], 1)
        b2, _ = np.polyfit(df['designY'], df['Y'], 1)
        rot = (a2 - b1) / 2
        v = {f'a0_{name}': a0, f'b0_{name}': b0,
           f'a1_{name}': a1, f'b1_{name}': b1,
           f'a2_{name}': a2, f'b2_{name}': b2,
           f'rot_{name}': rot}
        return v

# Excel処理部
    def to_excel(self, p_save):
        wb = openpyxl.load_workbook(self.p_excel)
        ws = wb['Data']

        # データフレーム処理部
        df1w = self.df1[['X1', 'Y1', 'X2', 'Y2']]
        df2w = self.df2[['X1', 'Y1', 'X2', 'Y2']]
        self._write_excel(df1w, ws)
        self._write_excel(df2w, ws, offset_columns=4)

        # 測定時間処理部
        ws['l4'] = self.meas_start_1st
        ws['l5'] = self.meas_start_2nd
        ws['m4'] = self.meas_end_1st
        ws['m5'] = self.meas_end_2nd

        # 保存
        wb.save(p_save)

    # データフレームをExcelに書き込むときに使用
    def _write_excel(self, df, ws, offset_columns=0):
        for j in range(len(df)):
            for i in range(len(df.columns)):
                ws.cell(row=self.excel_start_row+j, column=self.excel_start_column+offset_columns+i,
                        value=df.iloc[j, i])


if __name__ == '__main__':
    p_lms1 = Path('../../test/test_data/data_pos_qc_gpos/20231018QC_Global_v3000.000.lms')
    p_lms2 = Path('../../test/test_data/data_pos_qc_gpos/20231018QC_Global_v3000.001.lms')
    c = PosQcGpos(p_lms1, p_lms2)
    p_excel = Path('../../excel_template/Global STR format.xlsx')
    p_save = Path('../../../result/Global STR result.xlsx')
    c.to_excel(p_save)
    # print(vars(c))