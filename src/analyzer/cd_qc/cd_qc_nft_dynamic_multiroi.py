import os
import sys
import numpy as np
import pandas as pd
from pathlib import Path

import openpyxl

from pprint import pprint

sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
from getdf.convert_at import ConvertAt


class CdNftDynamicMultiRoi:
    # Excel関係
    p_excel = Path(os.path.dirname(__file__) + '/../../excel_template/cdQC/NFT Dynamic multiROI format.xlsx')
    excel_start_row = 4
    excel_start_column = 2

    def __init__(self, p_normal, p_no_overlap, plate_type=''):
        df_normal = self._convert_df(p_normal)
        df_no_overlap = self._convert_df(p_no_overlap)

        # 測定時間
        meas_start_normal = df_normal['meas_date'].min()
        meas_end_normal = df_normal['meas_date'].max()
        meas_start_no_overlap = df_no_overlap['meas_date'].min()
        meas_end_no_overlap = df_no_overlap['meas_date'].max()
        meas_time_normal = meas_end_normal - meas_start_normal
        meas_time_no_overlap = meas_end_no_overlap - meas_start_no_overlap
        v_meas = {'meas_time': meas_time_normal, 'meas_start': meas_start_normal, 'meas_end': meas_end_normal,
                  'meas_time_normal': meas_time_normal, 'meas_start_normal': meas_start_normal, 'meas_end_normal': meas_end_normal,
                  'meas_time_no_overlap': meas_time_no_overlap, 'meas_start_no_overlap': meas_start_no_overlap,
                  'meas_end_no_overlap': meas_end_no_overlap,
                  }

        # 値
        v_value = self._calc_angle(df_normal, df_no_overlap)

        # plate type
        self.plate_type = plate_type
        v_plate = {'plate_type': self.plate_type}

        # columnとrowをオブジェクト変数化
        self.column_num = df_normal['die_x'][0]
        self.row_num = df_normal['die_y'][0]

        # 結合してオブジェクト変数化
        v = v_meas | v_value | v_plate
        for k in v:
            setattr(self, k, v[k])

        # Excel用の変換
        self.df_normal = df_normal
        self.df_no_overlap = df_no_overlap

    def _convert_df(self, p):
        # strかpathlibかわからんのでとりあえず変換。
        p = Path(p)
        if not p:
            raise FileNotFoundError('ファイルがありません。')
        if p.stat().st_size == 0:
            raise FileNotFoundError('ファイルが空です。')

        # 変換
        c = ConvertAt()
        df = df = c.df(p)

        if df.empty:
            raise FileNotFoundError('ファイルが空です。')

        return df

    def _calc_angle(self, df_normal_org, df_no_overlap_org):
        # 深いコピーにして元df変更防止
        df_normal = df_normal_org.copy()
        df_no_overlap = df_no_overlap_org.copy()

        # df_normalの前処理。ついでにrefに名前変更。
        df_normal = df_normal[['design_pos_x_ROI1', 'design_pos_y_ROI1', 'label1', 'cd1']]
        df_normal = df_normal.rename(columns={'design_pos_x_ROI1': 'design_pos_x', 'design_pos_y_ROI1': 'design_pos_y',
                                              'label1': 'label', 'cd1': 'cd_ref'})
        df_ref_hor = df_normal[df_normal['label']=='SPACE-Y']
        df_ref_ver = df_normal[df_normal['label'] == 'SPACE-X']

        # df_no_overlapの前処理。ついでに名前を変更。
        list_columns = df_no_overlap.columns
        list_columns = list(filter(lambda x: x.startswith('cd'), list_columns))
        df = pd.DataFrame()
        for i, columns in enumerate(list_columns):
            idx = i + 1
            df_tmp = df_no_overlap[[f'design_pos_x_ROI{idx}', f'design_pos_y_ROI{idx}', f'label{idx}', f'cd{idx}']]
            df_tmp = df_tmp.rename(columns=
                        {f'design_pos_x_ROI{idx}': 'design_pos_x',
                         f'design_pos_y_ROI{idx}': 'design_pos_y',
                         f'label{idx}': 'label',
                         f'cd{idx}': 'cd', }
            )
            df = pd.concat([df, df_tmp])
        df_hor = df[df['label']=='SPACE-Y']
        df_ver = df[df['label']=='SPACE-X']

        # dfのマージ
        df_all_hor = pd.merge(df_hor, df_ref_hor, how='inner', on=['design_pos_x', 'design_pos_y'])
        df_all_ver = pd.merge(df_ver, df_ref_ver, how='inner', on=['design_pos_x', 'design_pos_y'])

        df_all_hor['cd_diff'] = df_all_hor['cd'] - df_all_hor['cd_ref']
        df_all_ver['cd_diff'] = df_all_ver['cd'] - df_all_ver['cd_ref']

        # 再現性
        rep_hor = df_all_hor['cd_diff'].std() * 3 / np.sqrt(2)
        rep_ver = df_all_ver['cd_diff'].std() * 3 / np.sqrt(2)

        # CD diff
        cd_diff_hor = df_all_hor['cd_diff'].mean()
        cd_diff_ver = df_all_ver['cd_diff'].mean()

        # cd mean
        cd_mean_hor = df_all_hor['cd'].mean()
        cd_mean_ref_hor = df_all_hor['cd_ref'].mean()
        cd_mean_ver = df_all_ver['cd'].mean()
        cd_mean_ref_ver = df_all_ver['cd_ref'].mean()

        # 3s
        cd_3s_hor = df_all_hor['cd'].std() * 3
        cd_3s_ref_hor = df_all_hor['cd_ref'].std() * 3
        cd_3s_ver = df_all_ver['cd'].std() * 3
        cd_3s_ref_ver = df_all_ver['cd_ref'].std() * 3

        # max
        cd_max_hor = df_all_hor['cd'].max()
        cd_max_ref_hor = df_all_hor['cd_ref'].max()
        cd_max_ver = df_all_ver['cd'].max()
        cd_max_ref_ver = df_all_ver['cd_ref'].max()

        # min
        cd_min_hor = df_all_hor['cd'].min()
        cd_min_ref_hor = df_all_hor['cd_ref'].min()
        cd_min_ver = df_all_ver['cd'].min()
        cd_min_ref_ver = df_all_ver['cd_ref'].min()

        # range
        cd_range_hor = df_all_hor['cd'].max()- df_all_hor['cd'].min()
        cd_range_ref_hor = df_all_hor['cd_ref'].max() - df_all_hor['cd_ref'].min()
        cd_range_ver = df_all_ver['cd'].max() - df_all_ver['cd'].min()
        cd_range_ref_ver = df_all_ver['cd_ref'].max() - df_all_ver['cd_ref'].min()

        v = {'rep_3s_hor': rep_hor, 'rep_3s_ver': rep_ver, 'cd_diff_hor': cd_diff_hor, 'cd_diff_ver': cd_diff_ver,
             'cd_mean_hor': cd_mean_hor, 'cd_mean_ref_hor': cd_mean_ref_hor,
             'cd_mean_ver': cd_mean_ver, 'cd_mean_ref_ver': cd_mean_ref_ver,
             'cd_3s_hor': cd_3s_hor, 'cd_3s_ref_hor': cd_3s_ref_hor,
             'cd_3s_ver': cd_3s_ver, 'cd_3s_ref_ver': cd_3s_ref_ver,
             'cd_max_hor': cd_max_hor, 'cd_max_ref_hor': cd_max_ref_hor,
             'cd_max_ver': cd_max_ver, 'cd_max_ref_ver': cd_max_ref_ver,
             'cd_min_hor': cd_min_hor, 'cd_min_ref_hor': cd_min_ref_hor,
             'cd_min_ver': cd_min_ver, 'cd_min_ref_ver': cd_min_ref_ver,
             'cd_range_hor': cd_range_hor, 'cd_range_ref_hor': cd_range_ref_hor,
             'cd_range_ver': cd_range_ver, 'cd_range_ref_ver': cd_range_ref_ver,
             }

        return v

# Excel処理部
    def to_excel(self, p_save):
        wb = openpyxl.load_workbook(self.p_excel)
        ws = wb['Data']

        # データフレーム処理部
        df_normal = self.df_normal[['cd1']]
        df_no_overlap = self.df_no_overlap[['cd1', 'cd2', 'cd3', 'cd4', 'cd5', 'cd6', 'cd7']]

        self._write_excel(df_normal, ws)
        self._write_excel(df_no_overlap, ws, offset_columns=1)

        # 測定時間処理部
        ws['o4'] = self.meas_start_normal
        ws['o5'] = self.meas_start_normal
        ws['o6'] = self.meas_end_normal
        ws['p4'] = self.meas_start_no_overlap
        ws['p5'] = self.meas_start_no_overlap
        ws['p6'] = self.meas_end_no_overlap

        # 測定チップ入力
        ws['l4'] = self.column_num
        ws['l5'] = self.row_num

        # plate名
        ws['l9'] = self.plate_type

        # 保存
        wb.save(p_save)

    # データフレームをExcelに書き込むときに使用
    def _write_excel(self, df, ws, offset_columns=0):
        for j in range(len(df)):
            for i in range(len(df.columns)):
                ws.cell(row=self.excel_start_row+j, column=self.excel_start_column+offset_columns+i,
                        value=df.iloc[j, i])


if __name__ == '__main__':
    p_normal = Path('../../test/test_data/data_cd_qc_nft_dynamic_multiroi/NFT_Dynamic_multiROI_Normal_V01_1-2.000H.csv')
    p_no_overlap = Path('../../test/test_data/data_cd_qc_nft_dynamic_multiroi/NFT_Dynamic_multiROI_No_overlap_V01_1-2.000H.csv')
    c = CdNftDynamicMultiRoi(p_normal, p_no_overlap, plate_type='No.1')
    p_save = Path('../../../result/nft dynamic multiroi.xlsx')
    print(vars(c))

    # c.to_excel(p_save)
